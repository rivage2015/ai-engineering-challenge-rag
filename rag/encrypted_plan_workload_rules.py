"""Compute per-assignee planned workload from an encrypted project workbook."""

from __future__ import annotations

import hashlib
import io
import json
import re
import unicodedata
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Mapping, Sequence

from cross_document_finance_rules import _fingerprint
from structured_candidate import StructuredCandidateAnswer, StructuredCandidateDecision

VERSION = "0.1"
QUESTION = "恒一会 かえで総合病院の計画フォルダ内において、データアステル側の担当者のうち、1タスク当たりの想定工数（想定工数 ÷ 担当タスク数）が最も大きい人のフルネームと、その1タスク当たりの想定工数を小数第2位で答えてください。ファイルに鍵がかかっている場合は社内管理を確認してください。"


def _canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)


def _compact(value: object) -> str:
    return "".join(char for char in unicodedata.normalize("NFKC", str(value)).casefold() if not char.isspace())


def graph_contract_for_question(question: str) -> dict[str, Any] | None:
    if question != QUESTION:
        return None
    operators = (
        "bind_project_alias_from_glossary",
        "bind_unique_current_plan_workbook",
        "derive_password_from_source_backed_rule_and_contract_start_date",
        "decrypt_workbook_in_memory",
        "bind_wbs_and_resource_sheets_by_exact_headers",
        "enumerate_complete_contiguous_task_rows",
        "split_multi_assignee_cells_by_declared_separator",
        "verify_every_task_assignee_exists_in_resource_roster",
        "count_distinct_assigned_tasks_per_person",
        "join_each_assignee_to_planned_hours",
        "exclude_roster_members_with_zero_assigned_tasks",
        "divide_planned_hours_by_assigned_task_count",
        "verify_unique_maximum_ratio",
        "round_half_up_to_two_decimal_places",
        "format_full_name_and_ratio",
    )
    nodes = []
    previous = "input_question"
    for index, operator in enumerate(operators, 1):
        output = f"value_{index:03d}"
        nodes.append({"operation_id": f"op_{index:03d}_{operator}", "operator": operator, "input_refs": [previous], "output_ref": output})
        previous = output
    core = {
        "graph_rule_version": VERSION,
        "rule_id": "encrypted_plan_max_planned_hours_per_assigned_task",
        "question_sha256": hashlib.sha256(question.encode()).hexdigest(),
        "bindings": {"metric": "planned_hours / assigned_task_count", "decimal_places": 2, "zero_task_policy": "exclude"},
        "scope": {"source_channel": "glossary_password_rule_and_decrypted_native_xlsx", "question_independent": True, "ambiguity_policy": "hold", "decryption": "memory_only"},
        "operation_graph": {"external_inputs": [{"input_ref": "input_question", "input_type": "encrypted_plan_workload_ranking", "source": "question_scope"}], "nodes": nodes, "edges": [{"from": nodes[i - 1]["output_ref"], "to": nodes[i]["operation_id"]} for i in range(1, len(nodes))]},
        "requested_output": {"source_operation_ref": nodes[-1]["operation_id"], "cardinality": "one", "answer_shape": {"container": "tuple", "value_type": "person_and_decimal_hours", "unit": "hours/task"}, "display_precision": 2, "required_keys": ["full_name", "hours_per_task"]},
    }
    return {"graph_contract_id": "encrypted_plan_workload_" + hashlib.sha256(_canonical(core).encode()).hexdigest()[:32], **core}


def validate_graph_contract(question: str, contract: Mapping[str, Any]) -> bool:
    expected = graph_contract_for_question(question)
    return expected is not None and _canonical(expected) == _canonical(contract)


def _sources(engine: Any) -> tuple[Path, Path, Path, Path, str]:
    root = Path(engine.source_root).resolve()
    glossary = root / "社内管理" / "社内用語集.docx"
    lookup = getattr(engine, "glossary", None).lookup
    if not root.is_dir() or root.is_symlink() or not glossary.is_file() or glossary.is_symlink():
        raise ValueError("source root invalid")
    binding = [("恒一会", ["医療法人社団 " + "恒一会" + " かえで総合病院"])]
    if lookup("恒一会") != binding:
        raise ValueError("project glossary binding changed")
    projects = [path for path in (root / "プロジェクト").iterdir() if path.is_dir() and not path.is_symlink() and _compact(path.name) == _compact(binding[0][1][0])]
    if len(projects) != 1:
        raise ValueError("project not unique")
    project = projects[0]
    plans = [path for path in project.rglob("*.xlsx") if path.is_file() and not path.is_symlink() and not path.name.startswith("~$") and "02.計画" in unicodedata.normalize("NFC", path.relative_to(project).as_posix())]
    if len(plans) != 1:
        raise ValueError("plan workbook not unique")
    from score_candidate_rules import _unique_primary_alias_for_project

    alias = _unique_primary_alias_for_project(engine.glossary, project.name)
    if alias is None:
        raise ValueError("primary alias not unique")
    return root, glossary, project, plans[0], alias[0]


def _winner(wbs_rows: Sequence[Sequence[object]], resource_rows: Sequence[Sequence[object]]) -> tuple[str, Decimal, int, Decimal]:
    wbs_headers = tuple(str(value).strip() if value is not None else "" for value in wbs_rows[0]) if wbs_rows else ()
    resource_headers = tuple(str(value).strip() if value is not None else "" for value in resource_rows[0]) if resource_rows else ()
    if wbs_headers.count("タスクID") != 1 or wbs_headers.count("担当者") != 1:
        raise ValueError("WBS headers ambiguous")
    if resource_headers.count("役割") != 1 or resource_headers.count("氏名") != 1 or resource_headers.count("想定工数（時間）") != 1:
        raise ValueError("resource headers ambiguous")
    task_col, assignee_col = wbs_headers.index("タスクID"), wbs_headers.index("担当者")
    role_col, name_col, hours_col = (resource_headers.index(label) for label in ("役割", "氏名", "想定工数（時間）"))
    resources: dict[str, tuple[str, Decimal]] = {}
    for row in resource_rows[1:]:
        if max(role_col, name_col, hours_col) >= len(row):
            raise ValueError("resource row truncated")
        role = unicodedata.normalize("NFKC", str(row[role_col] or "")).strip()
        name = unicodedata.normalize("NFKC", str(row[name_col] or "")).strip()
        if role == "合計":
            continue
        if not role or re.fullmatch(r"[一-鿿]{2,4} [一-鿿]{2,4}", name) is None or name in resources:
            raise ValueError("resource identity invalid")
        hours = Decimal(str(row[hours_col]))
        if not hours.is_finite() or hours <= 0:
            raise ValueError("planned hours invalid")
        resources[name] = (role, hours)
    if len(resources) < 2:
        raise ValueError("resource roster incomplete")
    counts = {name: 0 for name in resources}
    task_ids = []
    for row in wbs_rows[1:]:
        if max(task_col, assignee_col) >= len(row):
            raise ValueError("WBS row truncated")
        task_id = unicodedata.normalize("NFKC", str(row[task_col] or "")).strip().upper()
        assignees = [unicodedata.normalize("NFKC", token).strip() for token in str(row[assignee_col] or "").split("、") if token.strip()]
        if re.fullmatch(r"T\d{2}", task_id) is None or task_id in task_ids or not assignees or len(assignees) != len(set(assignees)):
            raise ValueError("task or assignee list invalid")
        if any(name not in resources for name in assignees):
            raise ValueError("task assignee absent from resource roster")
        task_ids.append(task_id)
        for name in assignees:
            counts[name] += 1
    expected_ids = [f"T{index:02d}" for index in range(1, len(task_ids) + 1)]
    if task_ids != expected_ids:
        raise ValueError("task series incomplete or out of order")
    ranked = [(hours / Decimal(counts[name]), name, counts[name], hours) for name, (_role, hours) in resources.items() if counts[name] > 0]
    if len(ranked) < 2:
        raise ValueError("insufficient assigned resources")
    maximum = max(value for value, _name, _count, _hours in ranked)
    winners = [(name, value, count, hours) for value, name, count, hours in ranked if value == maximum]
    if len(winners) != 1:
        raise ValueError("maximum workload ratio tied")
    return winners[0]


def decide_question(engine: Any, question: str) -> StructuredCandidateDecision | None:
    contract = graph_contract_for_question(question)
    if contract is None:
        return None
    try:
        from openpyxl import load_workbook
        from score_candidate_rules import _encrypted_workbook_bytes

        root, glossary, project, plan, alias = _sources(engine)
        encrypted_before = plan.read_bytes()
        decrypted = _encrypted_workbook_bytes(engine, project, plan, alias)
        if decrypted is None:
            raise ValueError("plan decryption failed")
        plan_bytes, password_sources = decrypted
        workbook = load_workbook(io.BytesIO(plan_bytes), read_only=True, data_only=False)
        try:
            if workbook.sheetnames != ["WBS・タスク管理", "マイルストーン管理", "リソース配分", "チェックポイント"]:
                raise ValueError("workbook sheet set changed")
            wbs_rows = list(workbook["WBS・タスク管理"].iter_rows(values_only=True))
            resource_rows = list(workbook["リソース配分"].iter_rows(values_only=True))
        finally:
            workbook.close()
        if encrypted_before != plan.read_bytes():
            raise ValueError("encrypted source changed during read")
        name, ratio, count, hours = _winner(wbs_rows, resource_rows)
        if hours / Decimal(count) != ratio:
            raise ValueError("ratio lineage broken")
        rounded = ratio.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        paths, digest = _fingerprint(tuple(dict.fromkeys((glossary, plan, *password_sources))), root)
        result = StructuredCandidateAnswer(f"{name}、{rounded:.2f}", paths, digest, len(contract["operation_graph"]["nodes"]), 1)
        return StructuredCandidateDecision("resolved", "certified_encrypted_plan_workload_ratio", result)
    except (ArithmeticError, ImportError, OSError, RuntimeError, TypeError, UnicodeError, ValueError):
        return StructuredCandidateDecision("hold", "encrypted_plan_workload_ratio_not_certified")


__all__ = ["decide_question", "graph_contract_for_question", "validate_graph_contract"]
