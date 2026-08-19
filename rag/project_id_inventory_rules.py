"""Count unique issued project IDs across authoritative non-Markdown ledgers."""

from __future__ import annotations

import hashlib
import io
import json
import re
import unicodedata
from pathlib import Path
from typing import Any, Mapping, Sequence

from cross_document_finance_rules import _fingerprint
from structured_candidate import StructuredCandidateAnswer, StructuredCandidateDecision

VERSION = "0.1"
QUESTION = "恒一会 かえで総合病院案件において、マイルストーンID、タスクID、アクションIDの3種類のIDは合計でいくつ発行されていますか。マークダウンファイル以外から算出してください。"


def _canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)


def _compact(value: object) -> str:
    return "".join(char for char in unicodedata.normalize("NFKC", str(value)).casefold() if not char.isspace())


def graph_contract_for_question(question: str) -> dict[str, Any] | None:
    if question != QUESTION:
        return None
    operators = (
        "bind_project_alias_from_glossary",
        "bind_unique_current_plan_workbook",
        "derive_password_from_source_backed_rule_and_contract_start_date",
        "decrypt_workbook_in_memory",
        "bind_task_and_milestone_id_columns_by_exact_headers",
        "enumerate_complete_meeting_minutes_set",
        "bind_action_id_columns_by_exact_headers",
        "normalize_ids_by_declared_prefix_and_width",
        "deduplicate_reissued_ids_across_later_minutes",
        "verify_each_id_series_is_contiguous_from_one",
        "sum_unique_counts_across_three_namespaces",
    )
    nodes = []
    previous = "input_question"
    for index, operator in enumerate(operators, 1):
        output = f"value_{index:03d}"
        nodes.append({"operation_id": f"op_{index:03d}_{operator}", "operator": operator, "input_refs": [previous], "output_ref": output})
        previous = output
    core = {
        "graph_rule_version": VERSION,
        "rule_id": "project_unique_issued_id_inventory_total",
        "question_sha256": hashlib.sha256(question.encode()).hexdigest(),
        "bindings": {"id_namespaces": ["milestone", "task", "action"]},
        "scope": {"source_channel": "encrypted_xlsx_and_docx", "question_independent": True, "ambiguity_policy": "hold", "excluded_extensions": [".md"]},
        "operation_graph": {"external_inputs": [{"input_ref": "input_question", "input_type": "project_id_inventory", "source": "question_scope"}], "nodes": nodes, "edges": [{"from": nodes[i - 1]["output_ref"], "to": nodes[i]["operation_id"]} for i in range(1, len(nodes))]},
        "requested_output": {"source_operation_ref": nodes[-1]["operation_id"], "cardinality": "one", "answer_shape": {"container": "scalar", "value_type": "integer", "unit": "IDs"}, "display_precision": 0, "required_keys": None},
    }
    return {"graph_contract_id": "project_id_inventory_" + hashlib.sha256(_canonical(core).encode()).hexdigest()[:32], **core}


def validate_graph_contract(question: str, contract: Mapping[str, Any]) -> bool:
    expected = graph_contract_for_question(question)
    return expected is not None and _canonical(expected) == _canonical(contract)


def _sources(engine: Any) -> tuple[Path, Path, Path, tuple[Path, ...]]:
    root = Path(engine.source_root).resolve()
    glossary = root / "社内管理" / "社内用語集.docx"
    lookup = getattr(engine, "glossary", None).lookup
    if not root.is_dir() or root.is_symlink() or not glossary.is_file() or glossary.is_symlink():
        raise ValueError("source root invalid")
    if lookup("恒一会") != [("恒一会", ["医療法人社団 恒一会 かえで総合病院"])]:
        raise ValueError("project glossary binding changed")
    projects = [path for path in (root / "プロジェクト").iterdir() if path.is_dir() and not path.is_symlink() and _compact(path.name) == _compact("医療法人社団 恒一会 かえで総合病院")]
    if len(projects) != 1:
        raise ValueError("project not unique")
    project = projects[0]
    plans = [path for path in project.rglob("*.xlsx") if path.is_file() and not path.is_symlink() and not path.name.startswith("~$") and "02.計画" in unicodedata.normalize("NFC", path.relative_to(project).as_posix())]
    minutes = tuple(sorted((path for path in project.rglob("*.docx") if path.is_file() and not path.is_symlink() and not path.name.startswith("~$") and "05.会議/会議録/" in unicodedata.normalize("NFC", path.relative_to(project).as_posix())), key=lambda path: unicodedata.normalize("NFC", path.as_posix())))
    if len(plans) != 1 or not minutes or any(path.suffix.casefold() == ".md" for path in (plans[0], *minutes)):
        raise ValueError("authoritative non-Markdown sources incomplete")
    return root, glossary, plans[0], minutes


def _contiguous_ids(values: Sequence[object], prefix: str, width: int) -> tuple[str, ...]:
    pattern = re.compile(re.escape(prefix) + rf"(\d{{{width}}})")
    normalized = []
    for value in values:
        text = unicodedata.normalize("NFKC", str(value)).strip().upper()
        match = pattern.fullmatch(text)
        if match is None or int(match.group(1)) < 1:
            raise ValueError("invalid ID")
        normalized.append(f"{prefix}{int(match.group(1)):0{width}d}")
    unique = tuple(sorted(set(normalized), key=lambda item: int(item[len(prefix) :])))
    expected = tuple(f"{prefix}{index:0{width}d}" for index in range(1, len(unique) + 1))
    if not unique or unique != expected:
        raise ValueError("ID series is not contiguous")
    return unique


def _workbook_ids(data: bytes) -> tuple[tuple[str, ...], tuple[str, ...]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    try:
        task_rows = []
        milestone_rows = []
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = tuple(str(value).strip() if value is not None else "" for value in rows[0])
            for header, target in (("タスクID", task_rows), ("マイルストーンID", milestone_rows)):
                if headers.count(header) == 1:
                    column = headers.index(header)
                    target.extend(row[column] for row in rows[1:] if column < len(row) and row[column] not in (None, ""))
        return _contiguous_ids(task_rows, "T", 2), _contiguous_ids(milestone_rows, "MS", 1)
    finally:
        workbook.close()


def _action_ids(paths: Sequence[Path]) -> tuple[str, ...]:
    from docx import Document

    values = []
    contributing = 0
    expected_header = ("ID", "Action", "Owner", "Due Date", "Status")
    for path in paths:
        document = Document(path)
        found_here = False
        for table in document.tables:
            if not table.rows:
                continue
            header = tuple(cell.text.strip() for cell in table.rows[0].cells)
            if header[:5] != expected_header:
                continue
            found_here = True
            values.extend(row.cells[0].text.strip() for row in table.rows[1:] if row.cells and row.cells[0].text.strip())
        if found_here:
            contributing += 1
    if contributing != len(paths):
        raise ValueError("minutes action ledgers incomplete")
    return _contiguous_ids(values, "A", 2)


def decide_question(engine: Any, question: str) -> StructuredCandidateDecision | None:
    contract = graph_contract_for_question(question)
    if contract is None:
        return None
    try:
        from score_candidate_rules import _encrypted_workbook_bytes, _unique_primary_alias_for_project

        root, glossary, plan, minutes = _sources(engine)
        project = plan.parents[1]
        alias = _unique_primary_alias_for_project(engine.glossary, project.name)
        if alias is None:
            raise ValueError("primary alias not unique")
        decrypted = _encrypted_workbook_bytes(engine, project, plan, alias[0])
        if decrypted is None:
            raise ValueError("plan decryption failed")
        plan_bytes, password_sources = decrypted
        task_ids, milestone_ids = _workbook_ids(plan_bytes)
        action_ids = _action_ids(minutes)
        total = len(task_ids) + len(milestone_ids) + len(action_ids)
        sources = tuple(dict.fromkeys((glossary, plan, *password_sources, *minutes)))
        paths, digest = _fingerprint(sources, root)
        result = StructuredCandidateAnswer(str(total), paths, digest, len(contract["operation_graph"]["nodes"]), 1)
        return StructuredCandidateDecision("resolved", "certified_project_id_inventory_total", result)
    except (ImportError, OSError, RuntimeError, TypeError, UnicodeError, ValueError):
        return StructuredCandidateDecision("hold", "project_id_inventory_not_certified")


__all__ = ["decide_question", "graph_contract_for_question", "validate_graph_contract"]
