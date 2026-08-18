"""Opt-in deterministic answers for certified tabular question families.

This module is deliberately separate from retrieval and answer generation.  It
uses only the raw question, a question-independent glossary, and source tables
under the shared-drive root.  If the question, source, table, or operation graph
is not uniquely determined, it returns a hold decision and the caller falls
back to the existing RAG path.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import json
import math
import re
import sys
import unicodedata
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from build_question_understanding import (  # noqa: E402
    build_question_understanding,
    derive_supported_intent_draft,
)
from structured_search_units import (  # noqa: E402
    DecodedTableRow,
    StructuredRowError,
    execute_operation_graph,
)


CANDIDATE_VERSION = "0.1"
SUPPORTED_SUFFIXES = frozenset({".csv", ".tsv", ".xlsx"})
MAX_SOURCE_BYTES = 256 * 1024 * 1024
MAX_ROWS = 1_000_000
MAX_HEADER_SCAN_ROWS = 100
_LIVE_GRAPH_CONTRACT_PREFIXES = (
    "docx_mixed_",
    "xlsx_highlight_",
    "xlsx_pivot_highlight_",
    "xlsx_histogram_",
    "xlsx_formula_ml_",
    "xlsx_version_diff_",
    "xlsx_role_task_",
    "notebook_version_diff_",
    "crossdoc_finance_",
    "cross_project_",
    "pptx_mixed_",
    "pptx_version_diff_",
    "pptx_spatial_",
    "pptx_schedule_",
    "pptx_revision_summary_",
    "pdfgraph_",
    "pdfrole_",
    "pdf_native_style_",
    "pdf_investment_coefficient_",
    "pdf_highlight_trend_",
    "pdf_action_transition_",
    "pdf_action_content_",
    "docx_page_structure_",
)
LEGAL_FORMS = (
    "株式会社",
    "有限会社",
    "合同会社",
    "合資会社",
    "合名会社",
    "医療法人社団",
    "医療法人財団",
    "医療法人",
    "一般社団法人",
    "一般財団法人",
    "公益社団法人",
    "公益財団法人",
    "学校法人",
    "社会福祉法人",
)
ORDINAL_PREFIX = re.compile(r"^\s*(?:第\s*)?[0-9０-９]+\s*[.．、:：)）・\-]\s*")


@dataclass(frozen=True)
class SourceTable:
    path: Path
    source_sha256: str
    table_name: str
    headers: tuple[str, ...]
    rows: tuple[tuple[str, ...], ...]
    table_sha256: str


@dataclass(frozen=True)
class StructuredCandidateAnswer:
    answer: str
    source_paths: tuple[str, ...]
    source_sha256: str
    operation_count: int
    output_count: int


@dataclass(frozen=True)
class StructuredCandidateDecision:
    status: str
    reason: str
    result: StructuredCandidateAnswer | None = None


def _normalized(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value)).casefold().strip()


def _legal_core(value: object) -> str:
    result = _normalized(value)
    for form in LEGAL_FORMS:
        result = result.replace(_normalized(form), "")
    return re.sub(r"\s+", "", result)


def _display_scalar(value: Any) -> str:
    if isinstance(value, Decimal):
        if not value.is_finite():
            raise StructuredRowError("structured output is non-finite")
        rendered = format(value, "f")
        if "." in rendered:
            rendered = rendered.rstrip("0").rstrip(".")
        return rendered or "0"
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError("table contains a non-finite number")
        return format(value, ".15g")
    return str(value).strip()


def _stable_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )


def _source_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _required_fields(intent: Mapping[str, Any]) -> tuple[str, ...]:
    fields: list[str] = []
    for node in intent["operation_graph"]["nodes"]:
        predicate = node.get("predicate")
        if isinstance(predicate, Mapping):
            fields.append(str(predicate["field"]))
        fields.extend(str(value) for value in (node.get("fields") or []))
        field = node.get("field")
        if isinstance(field, str):
            fields.append(field)
    return tuple(dict.fromkeys(fields))


def _table_from_matrix(
    path: Path,
    source_sha256: str,
    table_name: str,
    matrix: Sequence[Sequence[Any]],
    required_fields: Sequence[str],
) -> SourceTable | None:
    required = {_normalized(value): value for value in required_fields}
    if not required:
        return None
    header_index: int | None = None
    header_cells: tuple[str, ...] | None = None
    for index, row in enumerate(matrix[:MAX_HEADER_SCAN_ROWS]):
        cells = tuple(_cell_text(value) for value in row)
        normalized = [_normalized(value) for value in cells if value]
        if len(normalized) != len(set(normalized)):
            continue
        if set(required) <= set(normalized):
            header_index = index
            header_cells = cells
            break
    if header_index is None or header_cells is None:
        return None
    last = max(index for index, value in enumerate(header_cells) if value)
    headers = header_cells[: last + 1]
    if any(not value for value in headers):
        return None
    normalized_headers = tuple(_normalized(value) for value in headers)
    if len(normalized_headers) != len(set(normalized_headers)):
        return None
    required_indexes = {
        normalized_headers.index(normalized_field) for normalized_field in required
    }
    rows: list[tuple[str, ...]] = []
    for raw_row in matrix[header_index + 1 :]:
        values = tuple(
            _cell_text(raw_row[index]) if index < len(raw_row) else ""
            for index in range(len(headers))
        )
        if not any(values):
            continue
        if any(not values[index] for index in required_indexes):
            return None
        rows.append(values)
        if len(rows) > MAX_ROWS:
            raise ValueError("structured table exceeds the row limit")
    if not rows:
        return None
    core = {"headers": headers, "rows": rows}
    return SourceTable(
        path=path,
        source_sha256=source_sha256,
        table_name=table_name,
        headers=headers,
        rows=tuple(rows),
        table_sha256=hashlib.sha256(_stable_json(core).encode("utf-8")).hexdigest(),
    )


def _read_delimited(
    path: Path,
    source_sha256: str,
    required_fields: Sequence[str],
) -> tuple[SourceTable, ...]:
    raw = path.read_bytes()
    if len(raw) > MAX_SOURCE_BYTES:
        raise ValueError("structured source exceeds the byte limit")
    text: str | None = None
    for encoding in ("utf-8-sig", "cp932"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("structured source encoding is unsupported")
    delimiter = "\t" if path.suffix.casefold() == ".tsv" else ","
    matrix = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    table = _table_from_matrix(
        path,
        source_sha256,
        path.name,
        matrix,
        required_fields,
    )
    return () if table is None else (table,)


def _read_xlsx(
    path: Path,
    source_sha256: str,
    required_fields: Sequence[str],
) -> tuple[SourceTable, ...]:
    if path.stat().st_size > MAX_SOURCE_BYTES:
        raise ValueError("structured source exceeds the byte limit")
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: list[SourceTable] = []
    try:
        for worksheet in workbook.worksheets:
            matrix: list[tuple[Any, ...]] = []
            for row in worksheet.iter_rows(values_only=True):
                matrix.append(tuple(row))
                if len(matrix) > MAX_ROWS + MAX_HEADER_SCAN_ROWS:
                    raise ValueError("structured table exceeds the row limit")
            table = _table_from_matrix(
                path,
                source_sha256,
                worksheet.title,
                matrix,
                required_fields,
            )
            if table is not None:
                tables.append(table)
    finally:
        workbook.close()
    return tuple(tables)


def _read_source_tables(
    path: Path,
    required_fields: Sequence[str],
) -> tuple[SourceTable, ...]:
    source_sha256 = _source_sha256(path)
    suffix = path.suffix.casefold()
    if suffix in {".csv", ".tsv"}:
        return _read_delimited(path, source_sha256, required_fields)
    if suffix == ".xlsx":
        return _read_xlsx(path, source_sha256, required_fields)
    return ()


def _candidate_values(value: str, glossary: Any) -> tuple[str, ...]:
    values = [value]
    entries = getattr(glossary, "entries", {})
    for canonical in entries.get(value, []):
        if canonical not in values:
            values.append(str(canonical))
    return tuple(values)


def _location_matches(parts: Sequence[str], candidates: Sequence[str]) -> bool:
    normalized_parts = {_normalized(part) for part in parts}
    legal_parts = {_legal_core(part) for part in parts}
    return any(
        _normalized(candidate) in normalized_parts
        or (_legal_core(candidate) and _legal_core(candidate) in legal_parts)
        for candidate in candidates
    )


def _container_matches(path: Path, candidates: Sequence[str]) -> bool:
    names = {_normalized(path.name), _normalized(path.stem)}
    return any(_normalized(candidate) in names for candidate in candidates)


def _decorated_label(value: str) -> str:
    return _normalized(ORDINAL_PREFIX.sub("", value))


def _prepare_rows(
    table: SourceTable,
    intent: Mapping[str, Any],
) -> tuple[DecodedTableRow, ...]:
    header_index = {value: index for index, value in enumerate(table.headers)}
    eq_expected: dict[str, str] = {}
    for node in intent["operation_graph"]["nodes"]:
        predicate = node.get("predicate")
        if not isinstance(predicate, Mapping) or predicate["operator"] != "eq":
            continue
        expected = predicate["value"]
        if not isinstance(expected, str):
            continue
        field = str(predicate["field"])
        existing = eq_expected.get(field)
        if existing is not None and existing != expected:
            raise StructuredRowError("one field has incompatible equality predicates")
        eq_expected[field] = expected
    rows: list[DecodedTableRow] = []
    for row_number, raw_values in enumerate(table.rows, 1):
        values = list(raw_values)
        for field, expected in eq_expected.items():
            index = header_index[field]
            raw = values[index]
            if raw != expected and _decorated_label(raw) == _decorated_label(expected):
                values[index] = expected
        row_core = {
            "source_sha256": table.source_sha256,
            "table": table.table_name,
            "row_number": row_number,
            "values": values,
        }
        row_id = "su_direct_" + hashlib.sha256(
            _stable_json(row_core).encode("utf-8")
        ).hexdigest()[:32]
        rows.append(
            DecodedTableRow(
                row_id,
                "doc_direct_" + table.source_sha256[:32],
                table.headers,
                tuple(values),
            )
        )
    return tuple(rows)


def _output_label(
    output: Mapping[str, Any],
    nodes_by_id: Mapping[str, Mapping[str, Any]],
) -> str:
    node = nodes_by_id[output["source_operation_ref"]]
    fields = node.get("fields") or []
    if fields:
        return str(fields[0])
    if node["operator"] == "mean":
        return "平均値"
    return str(output["return_field"])


def _render_answer(
    requested: Mapping[str, Any],
    requested_outputs: Sequence[Mapping[str, Any]],
) -> str:
    nodes_by_id = {
        node["operation_id"]: node for node in requested["operation_graph"]["nodes"]
    }
    consumed_refs = {
        reference
        for node in requested["operation_graph"]["nodes"]
        for reference in node["input_refs"]
    }
    paired_outputs = list(zip(requested["requested_outputs"], requested_outputs))
    terminal_outputs = [
        pair
        for pair in paired_outputs
        if nodes_by_id[pair[0]["source_operation_ref"]]["output_ref"]
        not in consumed_refs
    ]
    # Aggregates that feed a later selection are proof intermediates.  Emit the
    # terminal requested value while retaining the aggregate in the operation
    # trace.  Independent requested outputs remain terminal and are all kept.
    if terminal_outputs:
        paired_outputs = terminal_outputs
    rendered: list[tuple[str, str]] = []
    for contract, output in paired_outputs:
        value = output["value"]
        if isinstance(value, tuple):
            text = "該当なし" if not value else "、".join(_display_scalar(item) for item in value)
        else:
            text = _display_scalar(value)
        rendered.append((_output_label(contract, nodes_by_id), text))
    if len(rendered) == 1:
        return rendered[0][1]
    return "、".join(f"{label}: {value}" for label, value in rendered)


class StructuredCandidateEngine:
    """Resolve and execute only fully certified question/source combinations."""

    def __init__(self, source_root: Path, glossary: Any) -> None:
        root = source_root.resolve()
        if not root.is_dir() or root.is_symlink():
            raise ValueError("structured source root must be a real directory")
        self.source_root = root
        self.glossary = glossary
        self.paths = tuple(
            sorted(
                (
                    path
                    for path in root.rglob("*")
                    if path.is_file()
                    and not path.is_symlink()
                    and not path.name.startswith("~$")
                    and path.suffix.casefold() in SUPPORTED_SUFFIXES
                ),
                key=lambda value: unicodedata.normalize(
                    "NFC", value.relative_to(root).as_posix()
                ),
            )
        )

    def _matching_paths(self, intent: Mapping[str, Any]) -> tuple[Path, ...]:
        scope = intent["scope"]
        location = scope.get("location")
        container = scope.get("container")
        if not isinstance(location, str) or not isinstance(container, str):
            return ()
        locations = _candidate_values(location, self.glossary)
        containers = _candidate_values(container, self.glossary)
        matches: list[Path] = []
        for path in self.paths:
            relative = path.relative_to(self.source_root)
            if not _location_matches(relative.parts[:-1], locations):
                continue
            if _container_matches(path, containers):
                matches.append(path)
        return tuple(matches)

    def _execute_intent(self, intent: Mapping[str, Any]) -> StructuredCandidateDecision:
        """Execute one already-built graph intent without compiling a question."""

        required = _required_fields(intent)
        matching_paths = self._matching_paths(intent)
        if not matching_paths:
            return StructuredCandidateDecision("hold", "source_not_unique")
        tables: list[SourceTable] = []
        for path in matching_paths:
            tables.extend(_read_source_tables(path, required))
        if not tables:
            return StructuredCandidateDecision("hold", "table_not_certified")
        by_digest: dict[str, list[SourceTable]] = {}
        for table in tables:
            by_digest.setdefault(table.table_sha256, []).append(table)
        if len(by_digest) != 1:
            return StructuredCandidateDecision("hold", "table_not_unique")
        duplicates = next(iter(by_digest.values()))
        table = min(
            duplicates,
            key=lambda item: (
                len(item.path.relative_to(self.source_root).parts),
                unicodedata.normalize(
                    "NFC", item.path.relative_to(self.source_root).as_posix()
                ),
                item.table_name,
            ),
        )
        rows = _prepare_rows(table, intent)
        execution = execute_operation_graph(intent, rows)
        answer = _render_answer(intent, execution.requested_outputs)
        if not answer.strip():
            return StructuredCandidateDecision("hold", "empty_output")
        source_paths = tuple(
            sorted(
                {
                    unicodedata.normalize(
                        "NFC", item.path.relative_to(self.source_root).as_posix()
                    )
                    for item in duplicates
                }
            )
        )
        result = StructuredCandidateAnswer(
            answer=answer,
            source_paths=source_paths,
            source_sha256=table.source_sha256,
            operation_count=len(execution.operation_values),
            output_count=len(execution.requested_outputs),
        )
        return StructuredCandidateDecision("resolved", "certified_graph", result)

    def decide_from_graph(
        self,
        question_id: str,
        question: str,
        graph_plan: Any,
    ) -> StructuredCandidateDecision:
        """Try structured execution only after the mandatory graph-planning step.

        Candidate branches are taken from the supplied, validated GraphPlan;
        this method never recompiles question understanding.  Extended
        deterministic rules remain available, but only after their own typed
        graph contract has been reconstructed and validated from the complete
        question.
        """

        del question_id
        try:
            # A complete extended grammar has already rebuilt and validated
            # its own typed graph.  Execute that source-specific graph before
            # the generic table reader: visual OOXML, nested JSON, and other
            # certified sources are not necessarily row tables, and trying to
            # coerce them first can either fail or resolve the wrong contract.
            from score_candidate_rules import (
                decide_extended,
                graph_contract_for_question,
                validate_graph_contract,
            )

            contract = graph_contract_for_question(question)
            if contract is not None and not validate_graph_contract(question, contract):
                return StructuredCandidateDecision("error", "extended_graph_invalid")
            if contract is not None:
                # Newly introduced native Office lanes require the exact live
                # GraphPlan that certified the complete question.  Do not let
                # their executors be called through a dummy or mismatched
                # plan.  Older extended lanes retain their legacy direct-test
                # compatibility until their fixtures are migrated.
                if str(contract.get("graph_contract_id", "")).startswith(
                    _LIVE_GRAPH_CONTRACT_PREFIXES
                ):
                    if (
                        graph_plan is None
                        or getattr(graph_plan, "original_question", None) != question
                        or getattr(graph_plan, "strict_status", None) != "pass"
                    ):
                        return StructuredCandidateDecision(
                            "hold", "extended_graph_plan_not_certified"
                        )
                    supplied_branches = getattr(graph_plan, "branch_intents", ())
                    if (
                        not isinstance(supplied_branches, tuple)
                        or len(supplied_branches) != 1
                    ):
                        return StructuredCandidateDecision(
                            "hold", "extended_graph_plan_not_certified"
                        )
                    supplied_branch = supplied_branches[0]
                    if (
                        not isinstance(supplied_branch, Mapping)
                        or supplied_branch.get("status") != "resolved"
                    ):
                        return StructuredCandidateDecision(
                            "hold", "extended_graph_plan_not_certified"
                        )
                    supplied_intent = (
                        supplied_branch.get("intent")
                        if isinstance(supplied_branch, Mapping)
                        else None
                    )
                    supplied_contract = (
                        supplied_intent.get("extended_graph_contract")
                        if isinstance(supplied_intent, Mapping)
                        else None
                    )
                    if supplied_contract != contract:
                        return StructuredCandidateDecision(
                            "hold", "extended_graph_plan_contract_mismatch"
                        )
                extended = decide_extended(self, "graph-runtime", question)
                if extended is not None:
                    return extended
                # A question that matched a certified extended grammar must
                # fail closed when its source-specific executor cannot prove
                # a result.  Falling through to the generic row reader would
                # let an unresolved chart/OOXML/JSON contract be answered from
                # an unrelated tabular interpretation of the same workbook.
                return StructuredCandidateDecision(
                    "hold", "extended_source_not_resolved"
                )

            raw_branches = getattr(graph_plan, "branch_intents", ())
            branches = [
                branch
                for branch in raw_branches
                if isinstance(branch, Mapping)
                and isinstance(branch.get("intent"), Mapping)
            ]
            if len(branches) == 1:
                intent = branches[0]["intent"]
                nodes = (intent.get("operation_graph") or {}).get("nodes") or []
                if any(
                    isinstance(node, Mapping) and node.get("operator") != "unknown"
                    for node in nodes
                ):
                    decision = self._execute_intent(intent)
                    if decision.status == "resolved":
                        return decision

            return StructuredCandidateDecision("hold", "graph_not_structured")
        except (OSError, UnicodeDecodeError, ValueError, StructuredRowError) as exc:
            return StructuredCandidateDecision(
                "error", f"{type(exc).__name__}: {str(exc)[:200]}"
            )

    def decide(self, question_id: str, question: str) -> StructuredCandidateDecision:
        question_input = {
            "question_id": str(question_id),
            "original_question": question,
        }
        try:
            if derive_supported_intent_draft(question_input) is None:
                # Extended rules are deliberately imported lazily: the rule
                # module reuses the immutable decision/result records here,
                # while this core engine remains usable without a cycle at
                # module-import time.
                from score_candidate_rules import (
                    decide_extended,
                    graph_contract_for_question,
                )

                contract = graph_contract_for_question(question)
                if (
                    contract is not None
                    and str(contract.get("graph_contract_id", "")).startswith(
                        _LIVE_GRAPH_CONTRACT_PREFIXES
                    )
                ):
                    return StructuredCandidateDecision(
                        "hold", "extended_graph_plan_required"
                    )

                extended = decide_extended(self, str(question_id), question)
                if extended is not None:
                    return extended
                return StructuredCandidateDecision("unsupported", "question_grammar")
            qur = build_question_understanding(question_input)
            if qur["final_status"] != "ready_for_retrieval":
                return StructuredCandidateDecision("hold", "intent_not_ready")
            branches = qur["candidate_query_paths"]
            if len(branches) != 1 or branches[0]["status"] not in {
                "pending",
                "resolved",
            }:
                return StructuredCandidateDecision("hold", "branch_not_unique")
            intent = branches[0]["candidate_intent"]
            decision = self._execute_intent(intent)
            if decision.status == "resolved" and decision.result is not None:
                return StructuredCandidateDecision(
                    "resolved", "certified", decision.result
                )
            return decision
        except (OSError, UnicodeDecodeError, ValueError, StructuredRowError) as exc:
            return StructuredCandidateDecision(
                "error", f"{type(exc).__name__}: {str(exc)[:200]}"
            )


__all__ = [
    "CANDIDATE_VERSION",
    "StructuredCandidateAnswer",
    "StructuredCandidateDecision",
    "StructuredCandidateEngine",
]
