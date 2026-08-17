"""Question-complete, fail-closed rules for native Excel semantics."""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Mapping, Sequence

from structured_candidate import (
    StructuredCandidateAnswer,
    StructuredCandidateDecision,
    _candidate_values,
    _location_matches,
)


EXCEL_NATIVE_RULE_VERSION = "0.1"

VISIBLE_HIGHLIGHT_CONDITION = re.compile(
    r"^(?P<location>.+?)の(?P<container>[^,、。]+?\.xlsx)において、"
    r"表示されている(?P<semantic>[^,、。]+?)シートで、"
    r"(?P<color>[^,、。]+?)ハイライトになっているセルの条件を"
    r"答えてください。?$"
)

REGRESSION_COEFFICIENT_APPLY = re.compile(
    r"^(?P<location>.+?)の(?P<container>[^,、。]+?\.xlsx)において、"
    r"回帰分析の結果として記載されている係数を"
    r"(?P<id_field>[A-Za-z_][A-Za-z0-9_]*)=(?P<id_value>[+-]?\d+(?:\.\d+)?)"
    r"のデータに当てはめたときの予測値はいくつですか。"
    r"小数第(?P<digits>[0-9０-９]+)位まで答えてください。?$"
)

_MAX_WORKBOOK_BYTES = 256 * 1024 * 1024
_SEMANTIC_SUFFIXES = ("シート", "sheet", "係数")
_YELLOW_NAMES = frozenset({"黄", "黄色", "yellow"})
_YELLOW_ARGB = "FFFFEB9C"


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )


def _normalized(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value)).casefold().strip()


def _semantic_core(value: object) -> str:
    result = re.sub(r"\s+", "", _normalized(value))
    changed = True
    while result and changed:
        changed = False
        for suffix in _SEMANTIC_SUFFIXES:
            if result.endswith(suffix) and len(result) > len(suffix):
                result = result[: -len(suffix)]
                changed = True
                break
    return result


def _field_key(value: object) -> str:
    return re.sub(r"\s+", " ", _normalized(value))


def _source_root(engine: Any) -> Path | None:
    try:
        root = Path(engine.source_root)
        if not root.is_dir() or root.is_symlink():
            return None
        return root.resolve()
    except (AttributeError, OSError, RuntimeError, TypeError):
        return None


def _named_workbooks(engine: Any, location: str, container: str) -> list[Path]:
    root = _source_root(engine)
    if root is None:
        return []
    locations = _candidate_values(location, getattr(engine, "glossary", None))
    names = {_normalized(value) for value in _candidate_values(container, getattr(engine, "glossary", None))}
    matches: list[Path] = []
    try:
        paths = root.rglob("*")
        for path in paths:
            if (
                not path.is_file()
                or path.is_symlink()
                or path.name.startswith(("~$", "."))
                or path.suffix.casefold() != ".xlsx"
                or _normalized(path.name) not in names
            ):
                continue
            relative = path.relative_to(root)
            if not _location_matches(relative.parts[:-1], locations):
                continue
            if path.stat().st_size <= 0 or path.stat().st_size > _MAX_WORKBOOK_BYTES:
                continue
            matches.append(path)
    except OSError:
        return []
    return sorted(set(matches), key=lambda item: item.as_posix())


def _nodes(operators: Sequence[str]) -> list[dict[str, Any]]:
    nodes: list[dict[str, Any]] = []
    previous = "input_question"
    for index, operator in enumerate(operators, 1):
        output = f"value_{index:03d}"
        nodes.append(
            {
                "operation_id": f"op_{index:03d}_{operator}",
                "operator": operator,
                "input_refs": [previous],
                "output_ref": output,
            }
        )
        previous = output
    return nodes


def _contract(
    question: str,
    rule_id: str,
    bindings: Mapping[str, Any],
    scope: Mapping[str, Any],
    operators: Sequence[str],
    value_type: str,
    display_precision: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    nodes = _nodes(operators)
    core = {
        "graph_rule_version": EXCEL_NATIVE_RULE_VERSION,
        "rule_id": rule_id,
        "question_sha256": hashlib.sha256(question.encode("utf-8")).hexdigest(),
        "bindings": dict(bindings),
        "scope": dict(scope),
        "operation_graph": {
            "external_inputs": [
                {
                    "input_ref": "input_question",
                    "input_type": "source_records",
                    "source": "question_scope",
                }
            ],
            "nodes": nodes,
            "edges": [
                {"from": nodes[index - 1]["output_ref"], "to": nodes[index]["operation_id"]}
                for index in range(1, len(nodes))
            ],
        },
        "requested_output": {
            "source_operation_ref": nodes[-1]["operation_id"],
            "cardinality": "single",
            "answer_shape": {
                "container": "scalar",
                "value_type": value_type,
                "unit": None,
            },
            "display_precision": dict(display_precision) if display_precision else None,
            "required_keys": None,
        },
    }
    return {
        "graph_contract_id": "excel_native_"
        + hashlib.sha256(_canonical_json(core).encode("utf-8")).hexdigest()[:32],
        **core,
    }


def graph_contract_for_question(question: str) -> dict[str, Any] | None:
    if not isinstance(question, str):
        return None
    highlight = VISIBLE_HIGHLIGHT_CONDITION.fullmatch(question)
    if highlight:
        bindings = {key: highlight[key] for key in ("location", "container", "semantic", "color")}
        return _contract(
            question,
            "excel_visible_conditional_format_predicate",
            bindings,
            {
                "location": bindings["location"],
                "container": bindings["container"],
                "sheet_state": "visible",
                "sheet_semantic": bindings["semantic"],
                "style_channel": "conditional_formatting_dxf_fill",
                "color": bindings["color"],
            },
            ("retrieve", "select_visible_sheet", "select_conditional_format", "verify_color", "project_predicate"),
            "string",
        )
    regression = REGRESSION_COEFFICIENT_APPLY.fullmatch(question)
    if regression:
        digits = int(unicodedata.normalize("NFKC", regression["digits"]))
        if not 0 <= digits <= 12:
            return None
        bindings = {
            key: regression[key]
            for key in ("location", "container", "id_field", "id_value", "digits")
        }
        return _contract(
            question,
            "excel_raw_regression_coefficient_prediction",
            bindings,
            {
                "location": bindings["location"],
                "container": bindings["container"],
                "coefficient_source": "unique_labeled_coefficient_table",
                "record_selector": {bindings["id_field"]: bindings["id_value"]},
                "feature_transform": "raw",
            },
            ("retrieve", "select_coefficient_table", "select_unique_record", "align_fields", "calculate", "round"),
            "number",
            {"mode": "decimal_places", "digits": digits},
        )
    return None


def validate_graph_contract(question: str, contract: Mapping[str, Any]) -> bool:
    expected = graph_contract_for_question(question)
    if expected is None or not isinstance(contract, Mapping):
        return False
    try:
        return _canonical_json(expected) == _canonical_json(contract)
    except (TypeError, ValueError):
        return False


def _argb_values(fill: Any) -> tuple[str, ...]:
    values: list[str] = []
    for color in (getattr(fill, "fgColor", None), getattr(fill, "bgColor", None)):
        value = getattr(color, "rgb", None)
        if isinstance(value, str) and re.fullmatch(r"[0-9A-Fa-f]{8}", value):
            values.append(value.upper())
    return tuple(dict.fromkeys(values))


def _is_declared_color(argb: str, declared: str) -> bool:
    return (
        _normalized(declared) in _YELLOW_NAMES
        and argb == _YELLOW_ARGB
    )


def _plain_decimal(value: object) -> Decimal | None:
    try:
        result = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        return None
    return result if result.is_finite() else None


_OPERATOR_TEXT = {
    "lessThan": "{field}が{value}未満",
    "lessThanOrEqual": "{field}が{value}以下",
    "greaterThan": "{field}が{value}より大きい",
    "greaterThanOrEqual": "{field}が{value}以上",
    "equal": "{field}が{value}に等しい",
    "notEqual": "{field}が{value}に等しくない",
}


def _visible_highlight_answer(workbook: Any, semantic: str, color: str) -> str | None:
    core = _semantic_core(semantic)
    sheets = [
        sheet
        for sheet in workbook.worksheets
        if sheet.sheet_state == "visible"
        and core
        and core == _semantic_core(sheet["A1"].value)
    ]
    if len(sheets) != 1:
        return None
    color_rules: list[Any] = []
    for conditional in sheets[0].conditional_formatting:
        for rule in conditional.rules:
            dxf = getattr(rule, "dxf", None)
            if dxf is None and isinstance(getattr(rule, "dxfId", None), int):
                styles = workbook._differential_styles.styles
                if not 0 <= rule.dxfId < len(styles):
                    continue
                dxf = styles[rule.dxfId]
            fill = getattr(dxf, "fill", None)
            if fill is None or not any(_is_declared_color(value, color) for value in _argb_values(fill)):
                continue
            color_rules.append(rule)
    if len(color_rules) != 1:
        return None
    rule = color_rules[0]
    if rule.type != "cellIs" or rule.operator not in _OPERATOR_TEXT:
        return None
    if not isinstance(rule.formula, list) or len(rule.formula) != 1:
        return None
    threshold = _plain_decimal(rule.formula[0])
    if threshold is None:
        return None
    rendered = format(threshold, "f")
    return _OPERATOR_TEXT[rule.operator].format(field=semantic, value=rendered)


_INTERCEPT_LABELS = frozenset({"切片", "intercept", "constant", "const"})


def _coefficient_tables(workbook: Any) -> list[tuple[Any, dict[str, Decimal], Decimal]]:
    tables: list[tuple[Any, dict[str, Decimal], Decimal]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if _normalized(cell.value) != _normalized("係数") or cell.column < 2:
                    continue
                label_column = cell.column - 1
                coefficients: dict[str, Decimal] = {}
                intercept: Decimal | None = None
                for row_number in range(cell.row + 1, sheet.max_row + 1):
                    label = sheet.cell(row_number, label_column).value
                    value = sheet.cell(row_number, cell.column).value
                    if label is None and value is None:
                        break
                    if not isinstance(label, str) or not label.strip():
                        coefficients = {}
                        intercept = None
                        break
                    number = _plain_decimal(value)
                    if number is None:
                        coefficients = {}
                        intercept = None
                        break
                    key = _field_key(label)
                    if key in _INTERCEPT_LABELS:
                        if intercept is not None:
                            coefficients = {}
                            intercept = None
                            break
                        intercept = number
                    elif key in coefficients:
                        coefficients = {}
                        intercept = None
                        break
                    else:
                        coefficients[key] = number
                if coefficients and intercept is not None:
                    tables.append((sheet, coefficients, intercept))
    return tables


def _numeric_equal(left: object, right: Decimal) -> bool:
    value = _plain_decimal(left)
    return value is not None and value == right


def _raw_regression_answer(
    workbook: Any,
    id_field: str,
    id_value: str,
    digits: int,
) -> str | None:
    tables = _coefficient_tables(workbook)
    if len(tables) != 1:
        return None
    _, coefficients, intercept = tables[0]
    target_id = _plain_decimal(id_value)
    if target_id is None:
        return None
    id_key = _field_key(id_field)
    if not id_key or id_key in coefficients:
        return None
    required = set(coefficients) | {id_key}
    tables: list[tuple[Any, int, dict[str, int] | None]] = []
    for sheet in workbook.worksheets:
        for header_row in range(1, min(sheet.max_row, 20) + 1):
            headers = [sheet.cell(header_row, column).value for column in range(1, sheet.max_column + 1)]
            normalized_headers = [
                _field_key(value) if value is not None else ""
                for value in headers
            ]
            if not required.issubset(set(normalized_headers)):
                continue
            nonempty = [value for value in normalized_headers if value]
            if len(nonempty) != len(set(nonempty)):
                tables.append((sheet, header_row, None))
                continue
            tables.append(
                (
                    sheet,
                    header_row,
                    {
                        value: index
                        for index, value in enumerate(normalized_headers)
                        if value
                    },
                )
            )
    if len(tables) != 1 or tables[0][2] is None:
        return None
    sheet, header_row, header_map = tables[0]
    assert header_map is not None
    rows: list[tuple[object, ...]] = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if _numeric_equal(values[header_map[id_key]], target_id):
            rows.append(values)
    if len(rows) != 1:
        return None
    record = rows[0]
    result = intercept
    for field, coefficient in coefficients.items():
        value = _plain_decimal(record[header_map[field]])
        if value is None:
            return None
        result += coefficient * value
    if not result.is_finite():
        return None
    quantum = Decimal(1).scaleb(-digits)
    rounded = result.quantize(quantum, rounding=ROUND_HALF_UP)
    return format(rounded, f".{digits}f")


def _source_digest(paths: Sequence[Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(set(paths), key=lambda item: item.as_posix()):
        digest.update(path.read_bytes())
    return digest.hexdigest()


def _resolved(answer: str, paths: Sequence[Path], root: Path, operations: int) -> StructuredCandidateDecision:
    return StructuredCandidateDecision(
        "resolved",
        "certified_excel_native",
        StructuredCandidateAnswer(
            answer=answer,
            source_paths=tuple(
                unicodedata.normalize("NFC", path.relative_to(root).as_posix())
                for path in sorted(set(paths), key=lambda item: item.as_posix())
            ),
            source_sha256=_source_digest(paths),
            operation_count=operations,
            output_count=1,
        ),
    )


def decide_question(engine: Any, question: str) -> StructuredCandidateDecision | None:
    contract = graph_contract_for_question(question)
    if contract is None:
        return None
    bindings = contract["bindings"]
    root = _source_root(engine)
    if root is None:
        return StructuredCandidateDecision("hold", "excel_source_root_invalid")
    paths = _named_workbooks(engine, bindings["location"], bindings["container"])
    if len(paths) != 1:
        return StructuredCandidateDecision("hold", "excel_workbook_not_unique")
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(paths[0], data_only=False, read_only=False)
        try:
            if contract["rule_id"] == "excel_visible_conditional_format_predicate":
                answer = _visible_highlight_answer(
                    workbook, bindings["semantic"], bindings["color"]
                )
            else:
                answer = _raw_regression_answer(
                    workbook,
                    bindings["id_field"],
                    bindings["id_value"],
                    int(unicodedata.normalize("NFKC", bindings["digits"])),
                )
        finally:
            workbook.close()
    except (OSError, ValueError, KeyError, TypeError, InvalidOperation):
        return StructuredCandidateDecision("hold", "excel_source_invalid")
    if answer is None:
        return StructuredCandidateDecision("hold", "excel_semantics_not_unique")
    return _resolved(
        answer,
        paths,
        root,
        len(contract["operation_graph"]["nodes"]),
    )


__all__ = [
    "EXCEL_NATIVE_RULE_VERSION",
    "decide_question",
    "graph_contract_for_question",
    "validate_graph_contract",
]
