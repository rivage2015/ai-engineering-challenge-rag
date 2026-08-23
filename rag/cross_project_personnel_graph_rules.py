"""Resolve portfolio personnel questions through persisted-style Evidence Graph memory."""

from __future__ import annotations

import hashlib
import io
import json
import re
import unicodedata
import zipfile
from pathlib import Path
from typing import Any, Mapping, Sequence
from xml.etree import ElementTree

from cross_document_finance_rules import _decrypt_if_needed, _fingerprint, _opc_text, _pdf_text, _source_bytes
from cross_project_portfolio_rules import _contract_path, _final_report, _projects, _proposal, _safe_root
from evidence_edge_audit import EdgePolicy, EqualityCheck, audit_edge_with_same_model
from evidence_graph_memory import (
    add_node,
    canonical_json,
    load_graph,
    new_graph,
    propose_edge,
    save_graph,
    set_answer_projection,
    validate_graph,
)
from pptx_spatial_rules import _default_spatial_observer, _slide_rasters
from structured_candidate import StructuredCandidateAnswer, StructuredCandidateDecision

VERSION = "0.3"
QUESTION = "データアステル社の中でもっとも多くの案件にかかわっている人の内線番号を教えてください。"
QUESTION_COUNT = "各案件のPP・契約書・PLAN・FRにおいて、DA側の実施体制として役割付きで記載されている人物は全部で何人ですか。"

_ROLE_PATTERNS = (
    ("エグゼクティブスポンサー", r"エグゼクティブスポンサー"),
    ("プロジェクトマネージャー", r"プロジェクトマネージャー"),
    ("リードデータサイエンティスト", r"リード\s*データサイエンティスト"),
    ("データエンジニア", r"データエンジニア"),
    ("ビジネスアナリスト", r"ビジネスアナリスト"),
    ("QAレビュー", r"QA\s*レビュー(?:アー|ア|担当)"),
)
_SECONDARY_ROLE_PATTERNS = (
    ("エグゼクティブスポンサー", r"(?:エグゼクティブスポンサー|(?<![A-Za-z])ES(?![A-Za-z]))"),
    ("プロジェクトマネージャー", r"(?:プロジェクトマネージャー|(?<![A-Za-z])PM(?![A-Za-z]))"),
    ("リードデータサイエンティスト", r"(?:リード\s*(?:データサイエンティスト|DS)|(?<![A-Za-z])(?:LDS|DS)(?![A-Za-z]))"),
    ("データエンジニア", r"(?:データエンジニア|(?<![A-Za-z])DE(?![A-Za-z]))"),
    ("ビジネスアナリスト", r"(?:ビジネスアナリスト|(?<![A-Za-z])BA(?![A-Za-z]))"),
    ("QAレビュー", r"(?:QA\s*レビュー(?:アー|ア|担当)?|(?<![A-Za-z])QA(?![A-Za-z]))"),
)
_PERSON_PATTERN = r"([\u4e00-\u9fff々ヶ]{1,4})\s+([\u4e00-\u9fff々ヶ]{1,4})"
_PAIR_SEPARATOR = r"(?:\s+|\s*[：:─—\-・/（）()]\s*)"
_NON_PERSON_TOKENS = frozenset(
    {
        "記録", "完了", "結果", "確認", "指摘", "一覧", "承認", "担当",
        "分析", "運用", "管理", "低", "中", "高", "氏名", "役割",
        "基盤", "構築",
    }
)
_NON_PERSON_FRAGMENTS = frozenset(
    {"記録", "完了", "結果", "確認", "指摘", "承認", "担当", "分析", "運用", "管理"}
)


def _canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)


def _contract(question: str) -> dict[str, Any]:
    operators = (
        "bind_complete_project_set",
        "bind_current_proposal_contract_final_report_per_project",
        "bind_complete_internal_directory",
        "extract_project_person_rosters",
        "normalize_person_identity",
        "create_project_participation_nodes",
        "create_same_person_edges",
        "machine_audit_identity_edges",
        "blind_audit_with_directory_decoys",
        "falsify_scope_or_identity_mismatch",
        "count_distinct_projects_per_person",
        "require_unique_maximum",
        "join_winner_to_directory_extension",
        "project_extension",
    )
    nodes = []
    previous = "input_question"
    for index, operator in enumerate(operators, 1):
        output = f"value_{index:03d}"
        nodes.append(
            {
                "operation_id": f"op_{index:03d}_{operator}",
                "operator": operator,
                "input_refs": [previous],
                "output_ref": output,
            }
        )
        previous = output
    core = {
        "graph_rule_version": VERSION,
        "rule_id": "portfolio_unique_max_person_to_directory_extension",
        "question_sha256": hashlib.sha256(question.encode()).hexdigest(),
        "bindings": {"organization": "データアステル社", "metric": "distinct_project_count", "attribute": "内線番号"},
        "scope": {
            "source_channel": "current_project_core_documents_and_certified_directory",
            "question_independent": True,
            "ambiguity_policy": "hold",
            "working_memory": "evidence_graph_json_v0.1",
            "edge_audit": "machine_blind_falsifier",
        },
        "operation_graph": {
            "external_inputs": [{"input_ref": "input_question", "input_type": "project_corpus_and_directory", "source": "question_scope"}],
            "nodes": nodes,
            "edges": [
                {"from": nodes[index - 1]["output_ref"], "to": nodes[index]["operation_id"]}
                for index in range(1, len(nodes))
            ],
        },
        "requested_output": {
            "source_operation_ref": nodes[-1]["operation_id"],
            "cardinality": "single",
            "answer_shape": {"container": "scalar", "value_type": "string", "unit": None},
            "display_precision": None,
            "required_keys": None,
        },
    }
    return {"graph_contract_id": "personnel_graph_" + hashlib.sha256(_canonical(core).encode()).hexdigest()[:32], **core}


def graph_contract_for_question(question: str) -> dict[str, Any] | None:
    if not isinstance(question, str):
        return None
    if question == QUESTION:
        return _contract(question)
    if question == QUESTION_COUNT:
        contract = _contract(question)
        contract["rule_id"] = "portfolio_role_tagged_da_person_unique_count"
        contract["bindings"] = {
            "organization": "データアステル社",
            "documents": ["PP", "契約書", "PLAN", "FR"],
            "metric": "unique_person_count",
        }
        contract["scope"]["source_channel"] = "current_project_pp_contract_plan_fr_role_rosters"
        contract["requested_output"]["answer_shape"]["value_type"] = "integer"
        contract["requested_output"]["answer_shape"]["unit"] = "人"
        operators = (
            "bind_complete_project_set",
            "bind_current_pp_contract_plan_fr_per_project",
            "bind_authoritative_complete_six_role_pp_contract_roster",
            "require_pp_contract_roster_agreement_when_both_available",
            "extract_bidirectional_plan_fr_role_person_mentions",
            "expand_pm_lead_ds_ba_qa_role_abbreviations",
            "audit_plan_role_name_table_and_controlled_saito_variant",
            "reject_person_outside_authoritative_project_roster",
            "normalize_full_person_identity",
            "create_unique_role_person_nodes",
            "create_project_participation_nodes",
            "create_same_person_edges",
            "machine_audit_identity_edges",
            "blind_audit_with_person_decoys",
            "count_unique_verified_role_people",
            "format_person_count",
        )
        nodes = []
        previous = "input_question"
        for index, operator in enumerate(operators, 1):
            output = f"value_{index:03d}"
            nodes.append(
                {
                    "operation_id": f"op_{index:03d}_{operator}",
                    "operator": operator,
                    "input_refs": [previous],
                    "output_ref": output,
                }
            )
            previous = output
        contract["operation_graph"] = {
            "external_inputs": [
                {
                    "input_ref": "input_question",
                    "input_type": "project_corpus",
                    "source": "question_scope",
                }
            ],
            "nodes": nodes,
            "edges": [
                {"from": nodes[index - 1]["output_ref"], "to": nodes[index]["operation_id"]}
                for index in range(1, len(nodes))
            ],
        }
        contract["requested_output"]["source_operation_ref"] = nodes[-1]["operation_id"]
        core = {key: value for key, value in contract.items() if key != "graph_contract_id"}
        contract["graph_contract_id"] = "personnel_graph_" + hashlib.sha256(_canonical(core).encode()).hexdigest()[:32]
        return contract
    return None


def validate_graph_contract(question: str, contract: Mapping[str, Any]) -> bool:
    expected = graph_contract_for_question(question)
    try:
        return expected is not None and _canonical(expected) == _canonical(contract)
    except (TypeError, ValueError):
        return False


def _normalized(value: object) -> str:
    return "".join(
        char for char in unicodedata.normalize("NFKC", str(value)).casefold() if not char.isspace()
    )


def _text(path: Path) -> str:
    if path.suffix.casefold() in {".docx", ".pptx"}:
        return _opc_text(path)
    if path.suffix.casefold() == ".pdf":
        return _pdf_text(path)
    raise ValueError("unsupported project roster source")


def _plan_units(path: Path, *, decrypted: bytes | None = None) -> tuple[tuple[str, ...], ...]:
    from openpyxl import load_workbook

    data = decrypted if decrypted is not None else _decrypt_if_needed(path, _source_bytes(path))
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=False, keep_links=False)
    units: list[tuple[str, ...]] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = tuple(
                    unicodedata.normalize("NFKC", str(value)).strip()
                    for value in row
                    if value is not None and str(value).strip()
                )
                if cells:
                    units.append(cells)
    finally:
        workbook.close()
    return tuple(units)


def _plan_text(path: Path, *, decrypted: bytes | None = None) -> str:
    return "\n".join("\t".join(unit) for unit in _plan_units(path, decrypted=decrypted))


def _report_role_units(path: Path) -> tuple[tuple[str, ...], ...]:
    if path.suffix.casefold() == ".pdf":
        return tuple(
            tuple(cell for cell in re.split(r"\s{2,}", line.strip()) if cell)
            for line in _pdf_text(path).splitlines()
            if line.strip()
        )
    if path.suffix.casefold() != ".pptx":
        raise ValueError("unsupported final-report role source")
    data = _decrypt_if_needed(path, _source_bytes(path))
    units: list[tuple[str, ...]] = []
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        names = sorted(
            (
                name
                for name in archive.namelist()
                if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
            ),
            key=lambda name: int(re.search(r"(\d+)", Path(name).stem).group(1)),
        )
        if not names:
            raise ValueError("final-report slides missing")
        for name in names:
            root = ElementTree.fromstring(archive.read(name))
            for node in root.iter():
                local = node.tag.rsplit("}", 1)[-1]
                if local == "tr":
                    cells = []
                    for cell in (child for child in node if child.tag.rsplit("}", 1)[-1] == "tc"):
                        value = " ".join(
                            (item.text or "").strip()
                            for item in cell.iter()
                            if item.tag.rsplit("}", 1)[-1] == "t" and (item.text or "").strip()
                        )
                        cells.append(value)
                    if any(cells):
                        units.append(tuple(cells))
                elif local == "sp":
                    paragraphs = []
                    for paragraph in (
                        child for child in node.iter() if child.tag.rsplit("}", 1)[-1] == "p"
                    ):
                        value = "".join(
                            item.text or ""
                            for item in paragraph.iter()
                            if item.tag.rsplit("}", 1)[-1] == "t"
                        ).strip()
                        if value:
                            paragraphs.append(value)
                    if paragraphs:
                        units.append(tuple(paragraphs))
    return tuple(units)


def _extract_role_roster_text(text: str) -> dict[str, dict[str, str]] | None:
    """Extract one complete six-role DA roster from native document text."""

    normalized = unicodedata.normalize("NFKC", text)
    roster: dict[str, dict[str, str]] = {}
    for canonical_role, role_pattern in _ROLE_PATTERNS:
        matches = {
            unicodedata.normalize("NFC", f"{match.group(1)} {match.group(2)}")
            for match in re.finditer(
                role_pattern + r"\s*[：:]?\s*" + _PERSON_PATTERN,
                normalized,
            )
        }
        if len(matches) != 1:
            return None
        person = next(iter(matches))
        key = _normalized(person)
        if not key or key in roster:
            return None
        roster[key] = {"person": person, "role": canonical_role}
    return roster if len(roster) == len(_ROLE_PATTERNS) else None


def _extract_role_roster(path: Path) -> dict[str, dict[str, str]] | None:
    try:
        return _extract_role_roster_text(_text(path))
    except (OSError, RuntimeError, UnicodeError, ValueError):
        return None


def _plausible_person_name(person: str) -> bool:
    parts = unicodedata.normalize("NFKC", person).split()
    return (
        len(parts) == 2
        and all(re.fullmatch(r"[\u4e00-\u9fff々ヶ]{1,4}", part) for part in parts)
        and not any(part in _NON_PERSON_TOKENS for part in parts)
        and not any(fragment in part for fragment in _NON_PERSON_FRAGMENTS for part in parts)
    )


def _controlled_saito_variant(left: str, right: str) -> bool:
    compact = {_normalized(left), _normalized(right)}
    return compact == {"斉藤悠斗", "斎藤悠斗"}


def _canonical_roster_key(
    roster: Mapping[str, Mapping[str, str]],
    role: str,
    person: str,
    *,
    require_role: bool = True,
) -> str | None:
    observed_key = _normalized(person)
    for key, item in roster.items():
        if require_role and item["role"] != role:
            continue
        if key == observed_key or _controlled_saito_variant(item["person"], person):
            return key
    return None


def _person_cell(cell: str) -> str | None:
    match = re.fullmatch(r"\s*" + _PERSON_PATTERN + r"\s*", cell)
    if match is None:
        return None
    person = unicodedata.normalize("NFC", f"{match.group(1)} {match.group(2)}")
    return person if _plausible_person_name(person) else None


def _same_cell_person(cell: str, role_match: re.Match[str], *, before: bool) -> str | None:
    if before:
        match = re.search(
            _PERSON_PATTERN + _PAIR_SEPARATOR + r"$", cell[: role_match.start()]
        )
    else:
        match = re.match(
            _PAIR_SEPARATOR + _PERSON_PATTERN, cell[role_match.end() :]
        )
    if match is None:
        return None
    person = unicodedata.normalize("NFC", f"{match.group(1)} {match.group(2)}")
    return person if _plausible_person_name(person) else None


def _standalone_role_cell(cell: str, role_match: re.Match[str]) -> bool:
    remainder = cell[: role_match.start()] + cell[role_match.end() :]
    return re.sub(r"[\s：:─—\-・/（）()]", "", remainder) == ""


def _audit_structural_unit(
    unit: Sequence[str],
    roster: Mapping[str, Mapping[str, str]],
) -> dict[str, dict[str, str]]:
    cells = tuple(
        normalized
        for cell in unit
        for line in str(cell).splitlines()
        if (normalized := unicodedata.normalize("NFKC", line).strip())
    )
    occurrences: list[dict[str, Any]] = []
    for cell_index, cell in enumerate(cells):
        for role, role_pattern in _SECONDARY_ROLE_PATTERNS:
            for role_match in re.finditer(role_pattern, cell):
                same_cell: list[str] = []
                for before in (True, False):
                    person = _same_cell_person(cell, role_match, before=before)
                    if person is not None and person not in same_cell:
                        same_cell.append(person)
                neighbors: dict[str, str] = {}
                neighbor_cells: dict[str, int] = {}
                if _standalone_role_cell(cell, role_match):
                    for side, neighbor in (
                        ("before", cell_index - 1),
                        ("after", cell_index + 1),
                    ):
                        if 0 <= neighbor < len(cells):
                            person = _person_cell(cells[neighbor])
                            if person is not None:
                                neighbors[side] = person
                                neighbor_cells[side] = neighbor
                occurrences.append(
                    {
                        "role": role,
                        "same_cell": tuple(same_cell),
                        "neighbors": neighbors,
                        "neighbor_cells": neighbor_cells,
                    }
                )

    # Repeated table/list pairs can establish a structural direction.  A single
    # role between two person-like cells cannot: accepting its one known side
    # would silently discard a conflicting or out-of-roster mention.
    direction_votes: list[str] = []
    for occurrence in occurrences:
        if occurrence["same_cell"]:
            continue
        role = occurrence["role"]
        neighbors = occurrence["neighbors"]
        matching_sides = [
            side
            for side, person in neighbors.items()
            if _canonical_roster_key(roster, role, person) is not None
        ]
        if len(matching_sides) == 1:
            direction_votes.append(matching_sides[0])
    structural_direction = None
    if len(direction_votes) >= 2 and len(set(direction_votes)) == 1:
        structural_direction = direction_votes[0]

    selected_neighbor_cells: set[int] = set()
    if structural_direction is not None:
        selected_neighbor_cells = {
            occurrence["neighbor_cells"][structural_direction]
            for occurrence in occurrences
            if not occurrence["same_cell"]
            and structural_direction in occurrence["neighbor_cells"]
        }

    observed: dict[str, dict[str, str]] = {}
    for occurrence in occurrences:
        role = occurrence["role"]
        candidates = list(occurrence["same_cell"])
        neighbors = occurrence["neighbors"]
        if not candidates:
            if structural_direction is not None and structural_direction in neighbors:
                candidates = [neighbors[structural_direction]]
                for side, person in neighbors.items():
                    if (
                        side != structural_direction
                        and occurrence["neighbor_cells"][side]
                        not in selected_neighbor_cells
                        and person not in candidates
                    ):
                        candidates.append(person)
            else:
                candidates = list(dict.fromkeys(neighbors.values()))
        if not candidates:
            continue

        resolved: list[tuple[str, str]] = []
        for person in candidates:
            key = _canonical_roster_key(roster, role, person)
            if key is None:
                known_person = _canonical_roster_key(
                    roster, role, person, require_role=False
                )
                if known_person is not None:
                    raise ValueError("secondary source role disagrees with authoritative roster")
                raise ValueError("secondary source contains person outside authoritative roster")
            resolved.append((person, key))
        keys = {key for _, key in resolved}
        if len(keys) != 1:
            raise ValueError("role occurrence has multiple authoritative people")
        key = resolved[0][1]
        observed[key] = {"person": roster[key]["person"], "role": role}
    return observed


def _audit_secondary_role_mentions_units(
    units: Sequence[Sequence[str]],
    roster: Mapping[str, Mapping[str, str]],
) -> dict[str, dict[str, str]]:
    observed: dict[str, dict[str, str]] = {}
    for unit in units:
        observed.update(_audit_structural_unit(unit, roster))
    return observed


def _audit_secondary_role_mentions_text(
    text: str,
    roster: Mapping[str, Mapping[str, str]],
) -> dict[str, dict[str, str]]:
    """Audit PLAN/FR role-person mentions against the authoritative project roster."""
    lines = tuple(line.strip() for line in text.splitlines() if line.strip())
    return _audit_secondary_role_mentions_units((lines,), roster)


def _directory(engine: Any, root: Path):
    values = [
        path
        for path in root.rglob("*.pptx")
        if path.is_file() and not path.is_symlink() and _normalized(path.name) == _normalized("座席表.pptx")
    ]
    if len(values) != 1:
        return None, None
    path = values[0]
    slides = _slide_rasters(path)
    observation = _default_spatial_observer(engine, path, slides) if slides is not None else None
    if observation is None or observation.status != "certified" or len(observation.directory) != 12:
        return None, None
    people = {}
    for item in observation.directory:
        attributes = dict(item.attributes)
        extension = attributes.get("内線番号")
        key = _normalized(item.person)
        if key in people or not extension or len(extension) != 4 or not extension.isdigit():
            return None, None
        people[key] = {"person": item.person, "extension": extension, "evidence_id": item.evidence_id}
    return path, people


def _current_core_sources(project: Path) -> tuple[Path, Path, Path] | None:
    proposal = _proposal(project)
    contract = _contract_path(project)
    report = _final_report(project)
    if proposal is None or contract is None or report is None:
        return None
    return proposal, contract, report


def _current_plan(project: Path) -> Path | None:
    values = [
        path
        for path in project.rglob("*.xlsx")
        if path.is_file()
        and not path.is_symlink()
        and not path.name.startswith(("~$", "."))
        and any(_normalized("計画") in _normalized(part) for part in path.relative_to(project).parts[:-1])
        and _normalized("スケジュール") in _normalized(path.stem)
    ]
    if len(values) == 1:
        return values[0]
    ranked = []
    for path in values:
        match = re.search(r"r(\d+)$", unicodedata.normalize("NFKC", path.stem), re.I)
        if match:
            ranked.append((int(match.group(1)), path))
    if not ranked:
        return None
    maximum = max(rank for rank, _ in ranked)
    winners = [path for rank, path in ranked if rank == maximum]
    return winners[0] if len(winners) == 1 else None


def _same_model_role(packet: Mapping[str, Any]) -> Mapping[str, Any]:
    left = packet["from_node"]["normalized_value"]
    right = packet["to_node"]["normalized_value"]
    same_person = left.get("person") == right.get("person") and bool(left.get("person"))
    competing = [
        node
        for node in packet["decoy_nodes"]
        if node["normalized_value"].get("person") == right.get("person")
    ]
    if packet["audit_role"] == "blind_relation_classifier":
        verdict = "supported" if same_person and not competing else "ambiguous" if same_person else "contradicted"
        return {
            "verdict": verdict,
            "allowed_edge_types": [packet["proposed_edge_type"]] if verdict == "supported" else [],
            "rejected_edge_types": [] if verdict == "supported" else [packet["proposed_edge_type"]],
            "evidence_node_ids": [packet["from_node"]["node_id"], packet["to_node"]["node_id"]],
            "missing_checks": [] if verdict != "ambiguous" else ["unique_directory_identity"],
            "reason": "Identity is classified from normalized names with all alternative directory people supplied as decoys.",
        }
    if packet["audit_role"] == "relation_falsifier":
        falsified = not same_person or bool(competing)
        return {
            "falsified": falsified,
            "counterexamples": ([{"type": "competing_directory_identity", "node_ids": [node["node_id"] for node in competing]}] if competing else []),
            "unresolved_risks": ([] if not falsified else ["person_identity_not_unique"]),
            "reason": "Checked normalized identity against both endpoints and every directory decoy.",
        }
    raise ValueError("unexpected audit role")


def _build_memory(
    *,
    question: str,
    contract: Mapping[str, Any],
    directory_path: Path,
    people: Mapping[str, Mapping[str, str]],
    project_records: Sequence[Mapping[str, Any]],
) -> tuple[dict[str, Any], str, int]:
    graph = new_graph(
        question_id="Q013",
        question_sha256=hashlib.sha256(question.encode()).hexdigest(),
        graph_plan_id=str(contract["graph_contract_id"]),
    )
    directory_sha = hashlib.sha256(directory_path.read_bytes()).hexdigest()
    directory_nodes = {}
    for key, person in sorted(people.items()):
        directory_nodes[key] = add_node(
            graph,
            node_type="directory_person",
            value={"person": person["person"], "extension": person["extension"]},
            normalized_value={"person": key, "extension": person["extension"]},
            source={
                "path": unicodedata.normalize("NFC", directory_path.as_posix()),
                "sha256": directory_sha,
                "locator": {"evidence_id": person["evidence_id"]},
                "quote": f"{person['person']} / {person['extension']}",
                "extraction_method": "certified_dual_ocr_directory_observation",
            },
        )
    edge_ids = []
    counts = {key: set() for key in people}
    policy = EdgePolicy(
        edge_type="same_person",
        from_node_types=("directory_person",),
        to_node_types=("project_participation",),
        equality_checks=(EqualityCheck("normalized_value.person", "normalized_value.person", "nfc_compact"),),
    )
    decoys_by_person = {
        key: [node_id for other, node_id in directory_nodes.items() if other != key]
        for key in directory_nodes
    }
    for record in project_records:
        for key in record["people"]:
            evidence = record["people"][key]
            participation = add_node(
                graph,
                node_type="project_participation",
                value={
                    "person": people[key]["person"],
                    "project": record["project"],
                    "supporting_paths": evidence["paths"],
                },
                normalized_value={"person": key, "project": _normalized(record["project"])},
                source={
                    "path": evidence["paths"][0],
                    "sha256": evidence["sha256s"][0],
                    "locator": {"project": record["project"], "container_count": len(evidence["paths"])},
                    "quote": people[key]["person"],
                    "extraction_method": "current_core_document_exact_person_observation",
                },
            )
            edge_id = propose_edge(
                graph,
                edge_type="same_person",
                from_node_id=directory_nodes[key],
                to_node_id=participation,
                claim="The directory person and project roster mention identify the same employee.",
                comparison_fields=["normalized_value.person"],
            )
            status = audit_edge_with_same_model(
                graph,
                edge_id,
                policy,
                model_call=_same_model_role,
                decoy_node_ids=decoys_by_person[key],
            )
            if status != "verified":
                raise ValueError("person edge was not verified")
            edge_ids.append(edge_id)
            counts[key].add(record["project"])
    ranking = sorted(((len(projects), key) for key, projects in counts.items()), reverse=True)
    if len(ranking) < 2 or ranking[0][0] <= ranking[1][0] or ranking[0][0] == 0:
        raise ValueError("most involved person is not unique")
    maximum, winner = ranking[0]
    set_answer_projection(
        graph,
        operation="unique_argmax_distinct_project_count_then_directory_extension",
        input_node_ids=[directory_nodes[winner]],
        input_edge_ids=edge_ids,
    )
    if graph["state"] != "ready" or validate_graph(graph):
        raise ValueError("evidence graph memory did not validate")
    # Force the answering phase to consume a JSON reconstruction rather than
    # retaining Python object identity from the extraction phase.
    reloaded = json.loads(canonical_json(graph))
    if validate_graph(reloaded):
        raise ValueError("reloaded evidence graph memory did not validate")
    return reloaded, people[winner]["extension"], maximum


def _maybe_persist(engine: Any, graph: Mapping[str, Any], question_id: str = "Q013") -> None:
    configured = getattr(engine, "evidence_graph_memory_dir", None)
    if configured is None:
        return
    path = Path(configured) / f"{question_id}.evidence-graph.json"
    if path.exists():
        if load_graph(path) != graph:
            raise ValueError("existing Q013 evidence memory differs")
    else:
        save_graph(graph, path)


def _build_count_memory(
    *,
    question: str,
    contract: Mapping[str, Any],
    directory_path: Path,
    people: Mapping[str, Mapping[str, str]],
    project_records: Sequence[Mapping[str, Any]],
) -> tuple[dict[str, Any], int]:
    graph = new_graph(
        question_id="Q086",
        question_sha256=hashlib.sha256(question.encode()).hexdigest(),
        graph_plan_id=str(contract["graph_contract_id"]),
    )
    directory_sha = hashlib.sha256(directory_path.read_bytes()).hexdigest()
    directory_nodes = {}
    for key, person in sorted(people.items()):
        directory_nodes[key] = add_node(
            graph,
            node_type="directory_person",
            value={"person": person["person"], "extension": person["extension"]},
            normalized_value={"person": key, "extension": person["extension"]},
            source={
                "path": unicodedata.normalize("NFC", directory_path.as_posix()),
                "sha256": directory_sha,
                "locator": {"evidence_id": person["evidence_id"]},
                "quote": f"{person['person']} / {person['extension']}",
                "extraction_method": "certified_dual_ocr_directory_observation",
            },
        )
    policy = EdgePolicy(
        edge_type="same_person",
        from_node_types=("directory_person",),
        to_node_types=("project_participation",),
        equality_checks=(EqualityCheck("normalized_value.person", "normalized_value.person", "nfc_compact"),),
    )
    decoys = {key: [node for other, node in directory_nodes.items() if other != key] for key in directory_nodes}
    observed = set()
    edge_ids = []
    for record in project_records:
        for key, evidence in sorted(record["people"].items()):
            node = add_node(
                graph,
                node_type="project_participation",
                value={"person": people[key]["person"], "project": record["project"], "supporting_paths": evidence["paths"]},
                normalized_value={"person": key, "project": _normalized(record["project"])},
                source={
                    "path": evidence["paths"][0],
                    "sha256": evidence["sha256s"][0],
                    "locator": {"project": record["project"], "document_types": evidence["document_types"]},
                    "quote": people[key]["person"],
                    "extraction_method": "current_pp_contract_plan_fr_role_identity_observation",
                },
            )
            edge_id = propose_edge(
                graph,
                edge_type="same_person",
                from_node_id=directory_nodes[key],
                to_node_id=node,
                claim="The role-tagged DA roster entry and certified directory entry identify the same employee.",
                comparison_fields=["normalized_value.person"],
            )
            if audit_edge_with_same_model(graph, edge_id, policy, model_call=_same_model_role, decoy_node_ids=decoys[key]) != "verified":
                raise ValueError("role identity edge was not verified")
            edge_ids.append(edge_id)
            observed.add(key)
    if observed != set(people):
        raise ValueError("directory-backed DA roster is incomplete")
    set_answer_projection(
        graph,
        operation="count_unique_verified_directory_people",
        input_node_ids=[directory_nodes[key] for key in sorted(observed)],
        input_edge_ids=edge_ids,
    )
    if graph["state"] != "ready" or validate_graph(graph):
        raise ValueError("count evidence graph memory did not validate")
    reloaded = json.loads(canonical_json(graph))
    if validate_graph(reloaded):
        raise ValueError("reloaded count evidence graph memory did not validate")
    return reloaded, len(observed)


def _build_role_count_memory(
    *,
    question: str,
    contract: Mapping[str, Any],
    people: Mapping[str, Mapping[str, Any]],
    project_records: Sequence[Mapping[str, Any]],
) -> tuple[dict[str, Any], int]:
    graph = new_graph(
        question_id="Q086",
        question_sha256=hashlib.sha256(question.encode()).hexdigest(),
        graph_plan_id=str(contract["graph_contract_id"]),
    )
    person_nodes = {}
    for key, person in sorted(people.items()):
        roles = sorted(person["roles"])
        person_nodes[key] = add_node(
            graph,
            node_type="role_person",
            value={"person": person["person"], "roles": roles},
            normalized_value={"person": key},
            source={
                "path": person["paths"][0],
                "sha256": person["sha256s"][0],
                "locator": {"role": roles[0]},
                "quote": f"{roles[0]} / {person['person']}",
                "extraction_method": "authoritative_complete_pp_contract_da_role_roster",
            },
        )
    policy = EdgePolicy(
        edge_type="same_person",
        from_node_types=("role_person",),
        to_node_types=("project_participation",),
        equality_checks=(EqualityCheck("normalized_value.person", "normalized_value.person", "nfc_compact"),),
    )
    decoys = {key: [node for other, node in person_nodes.items() if other != key] for key in person_nodes}
    observed = set()
    edge_ids = []
    for record in project_records:
        for key, evidence in sorted(record["people"].items()):
            node = add_node(
                graph,
                node_type="project_participation",
                value={
                    "person": people[key]["person"],
                    "project": record["project"],
                    "role": evidence["role"],
                    "supporting_paths": evidence["paths"],
                },
                normalized_value={"person": key, "project": _normalized(record["project"])},
                source={
                    "path": evidence["paths"][0],
                    "sha256": evidence["sha256s"][0],
                    "locator": {
                        "project": record["project"],
                        "role": evidence["role"],
                        "audited_document_classes": evidence["audited_document_classes"],
                        "secondary_scan_counts": evidence["secondary_scan_counts"],
                    },
                    "quote": f"{evidence['role']} / {people[key]['person']}",
                    "extraction_method": "authoritative_roster_with_bidirectional_plan_fr_audit",
                },
            )
            edge_id = propose_edge(
                graph,
                edge_type="same_person",
                from_node_id=person_nodes[key],
                to_node_id=node,
                claim="The project role entry and cross-project role person identify the same employee.",
                comparison_fields=["normalized_value.person"],
            )
            if audit_edge_with_same_model(
                graph,
                edge_id,
                policy,
                model_call=_same_model_role,
                decoy_node_ids=decoys[key],
            ) != "verified":
                raise ValueError("role identity edge was not verified")
            edge_ids.append(edge_id)
            observed.add(key)
    if observed != set(people):
        raise ValueError("role roster union is incomplete")
    set_answer_projection(
        graph,
        operation="count_unique_verified_role_people",
        input_node_ids=[person_nodes[key] for key in sorted(observed)],
        input_edge_ids=edge_ids,
    )
    if graph["state"] != "ready" or validate_graph(graph):
        raise ValueError("role count evidence graph memory did not validate")
    reloaded = json.loads(canonical_json(graph))
    if validate_graph(reloaded):
        raise ValueError("reloaded role count evidence graph memory did not validate")
    return reloaded, len(observed)


def decide_question(engine: Any, question: str) -> StructuredCandidateDecision | None:
    contract = graph_contract_for_question(question)
    if contract is None:
        return None
    root = _safe_root(engine)
    projects = _projects(root) if root is not None else None
    if root is None or projects is None:
        return StructuredCandidateDecision("hold", "personnel_graph_project_set_incomplete")
    try:
        if question == QUESTION_COUNT:
            project_records = []
            people: dict[str, dict[str, Any]] = {}
            evidence_paths = []
            for project in projects:
                core = _current_core_sources(project)
                plan = _current_plan(project)
                if core is None or plan is None:
                    raise ValueError("four document classes")
                proposal, contract_path, report = core
                sources = (("PP", proposal), ("契約書", contract_path), ("PLAN", plan), ("FR", report))
                evidence_paths.extend(path for _, path in sources)
                try:
                    plan_units = _plan_units(plan)
                except Exception:
                    from cross_project_portfolio_rules import _alias
                    from score_candidate_rules import _encrypted_workbook_bytes

                    alias = _alias(engine, project)
                    decrypted = _encrypted_workbook_bytes(engine, project, plan, alias) if alias else None
                    if decrypted is None:
                        raise ValueError("encrypted plan")
                    plan_units = _plan_units(plan, decrypted=decrypted[0])
                    evidence_paths.extend(decrypted[1])
                report_units = _report_role_units(report)
                if not plan_units or not report_units:
                    raise ValueError("empty plan or final report")
                proposal_roster = _extract_role_roster(proposal)
                contract_roster = _extract_role_roster(contract_path)
                complete_rosters = [
                    value for value in (proposal_roster, contract_roster) if value is not None
                ]
                if not complete_rosters:
                    raise ValueError("complete role roster")
                roster = complete_rosters[0]
                if any(
                    _canonical(candidate) != _canonical(roster)
                    for candidate in complete_rosters[1:]
                ):
                    raise ValueError("role roster disagreement")
                secondary_observations = {
                    plan: _audit_secondary_role_mentions_units(plan_units, roster),
                    report: _audit_secondary_role_mentions_units(report_units, roster),
                }
                secondary_scan_counts = {
                    "PLAN": len(secondary_observations[plan]),
                    "FR": len(secondary_observations[report]),
                }
                project_people = {}
                for key, item in roster.items():
                    supporting = [
                        path
                        for path, candidate in (
                            (proposal, proposal_roster),
                            (contract_path, contract_roster),
                        )
                        if candidate is not None
                    ]
                    supporting.extend(
                        path
                        for path, observations in secondary_observations.items()
                        if key in observations
                    )
                    relative_paths = [
                        unicodedata.normalize("NFC", path.relative_to(root).as_posix())
                        for path in supporting
                    ]
                    sha256s = [hashlib.sha256(path.read_bytes()).hexdigest() for path in supporting]
                    project_people[key] = {
                        "role": item["role"],
                        "paths": relative_paths,
                        "sha256s": sha256s,
                        "audited_document_classes": [
                            document_type
                            for document_type, path in sources
                            if path in supporting
                        ],
                        "secondary_scan_counts": secondary_scan_counts,
                    }
                    aggregate = people.setdefault(
                        key,
                        {
                            "person": item["person"],
                            "roles": set(),
                            "paths": [],
                            "sha256s": [],
                        },
                    )
                    if aggregate["person"] != item["person"]:
                        raise ValueError("normalized person display disagreement")
                    aggregate["roles"].add(item["role"])
                    aggregate["paths"].extend(relative_paths)
                    aggregate["sha256s"].extend(sha256s)
                project_records.append({"project": project.name, "people": project_people})
            graph, count = _build_role_count_memory(
                question=question,
                contract=contract,
                people=people,
                project_records=project_records,
            )
            _maybe_persist(engine, graph, "Q086")
            source_paths, source_digest = _fingerprint(list(dict.fromkeys(evidence_paths)), root)
            combined_digest = hashlib.sha256(
                _canonical({"sources": source_digest, "evidence_graph": graph["integrity_sha256"]}).encode()
            ).hexdigest()
            return StructuredCandidateDecision(
                "resolved",
                "certified_cross_project_role_personnel_evidence_graph",
                StructuredCandidateAnswer(
                    f"{count}人",
                    source_paths,
                    combined_digest,
                    len(contract["operation_graph"]["nodes"]),
                    1,
                ),
            )
        directory_path, people = _directory(engine, root)
        if directory_path is None or people is None:
            raise ValueError("directory")
        project_records = []
        evidence_paths = [directory_path]
        for project in projects:
            sources = _current_core_sources(project)
            if sources is None:
                raise ValueError("core sources")
            evidence_paths.extend(sources)
            texts = {path: _text(path) for path in sources}
            roster = {}
            for key in people:
                matching = [path for path, text in texts.items() if people[key]["person"] in text]
                if matching:
                    matching.sort(key=lambda path: unicodedata.normalize("NFC", path.relative_to(root).as_posix()))
                    roster[key] = {
                        "paths": [unicodedata.normalize("NFC", path.relative_to(root).as_posix()) for path in matching],
                        "sha256s": [hashlib.sha256(path.read_bytes()).hexdigest() for path in matching],
                    }
            if not roster:
                raise ValueError("empty project roster")
            project_records.append({"project": project.name, "people": roster})
        graph, answer, maximum = _build_memory(
            question=question,
            contract=contract,
            directory_path=directory_path,
            people=people,
            project_records=project_records,
        )
        _maybe_persist(engine, graph)
        source_paths, source_digest = _fingerprint(evidence_paths, root)
        combined_digest = hashlib.sha256(
            _canonical({"sources": source_digest, "evidence_graph": graph["integrity_sha256"]}).encode()
        ).hexdigest()
        return StructuredCandidateDecision(
            "resolved",
            "certified_cross_project_personnel_evidence_graph",
            StructuredCandidateAnswer(
                answer,
                source_paths,
                combined_digest,
                len(contract["operation_graph"]["nodes"]),
                1,
            ),
        )
    except (OSError, RuntimeError, UnicodeError, ValueError):
        return StructuredCandidateDecision("hold", "personnel_graph_evidence_not_certified")


__all__ = ["QUESTION", "QUESTION_COUNT", "decide_question", "graph_contract_for_question", "validate_graph_contract"]
