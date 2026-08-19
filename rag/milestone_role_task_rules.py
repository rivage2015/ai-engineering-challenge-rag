"""Join a presentation role roster to milestone-linked workbook task rows."""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
import zipfile
from pathlib import Path
from typing import Any, Mapping, Sequence
from xml.etree import ElementTree as ET

from cross_document_finance_rules import _fingerprint
from structured_candidate import StructuredCandidateAnswer, StructuredCandidateDecision

VERSION = "0.1"
QUESTION = "蒼樹会 みなみ野女性医療センターのスケジュール.xlsxにおいて、MS3に紐づくタスクのうち、ビジネスアナリストが関わっているタスクIDを答えてください。"
_NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
}


def _canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)


def _compact(value: object) -> str:
    return "".join(char for char in unicodedata.normalize("NFKC", str(value)).casefold() if not char.isspace())


def graph_contract_for_question(question: str) -> dict[str, Any] | None:
    if question != QUESTION:
        return None
    operators = (
        "bind_project_alias_from_glossary",
        "bind_unique_current_proposal_and_named_schedule",
        "extract_role_person_edge_from_native_presentation_roster",
        "bind_schedule_header_semantics",
        "enumerate_complete_task_rows",
        "filter_rows_by_exact_milestone_token",
        "filter_rows_by_exact_assigned_person_token",
        "project_task_ids",
        "verify_task_ids_are_unique_and_well_formed",
        "return_all_matching_task_ids_in_source_order",
    )
    nodes = []
    previous = "input_question"
    for index, operator in enumerate(operators, 1):
        output = f"value_{index:03d}"
        nodes.append({"operation_id": f"op_{index:03d}_{operator}", "operator": operator, "input_refs": [previous], "output_ref": output})
        previous = output
    core = {
        "graph_rule_version": VERSION,
        "rule_id": "milestone_and_role_to_task_ids",
        "question_sha256": hashlib.sha256(question.encode()).hexdigest(),
        "bindings": {"milestone_id": "MS3", "role": "ビジネスアナリスト"},
        "scope": {"source_channel": "native_pptx_roster_and_xlsx_task_rows", "question_independent": True, "ambiguity_policy": "hold"},
        "operation_graph": {"external_inputs": [{"input_ref": "input_question", "input_type": "role_milestone_task_join", "source": "question_scope"}], "nodes": nodes, "edges": [{"from": nodes[i - 1]["output_ref"], "to": nodes[i]["operation_id"]} for i in range(1, len(nodes))]},
        "requested_output": {"source_operation_ref": nodes[-1]["operation_id"], "cardinality": "multiple", "answer_shape": {"container": "list", "value_type": "task_id", "unit": None}, "display_precision": None, "required_keys": None},
    }
    return {"graph_contract_id": "milestone_role_task_" + hashlib.sha256(_canonical(core).encode()).hexdigest()[:32], **core}


def validate_graph_contract(question: str, contract: Mapping[str, Any]) -> bool:
    expected = graph_contract_for_question(question)
    return expected is not None and _canonical(expected) == _canonical(contract)


def _sources(engine: Any) -> tuple[Path, Path, Path, Path]:
    root = Path(engine.source_root).resolve()
    glossary = root / "社内管理" / "社内用語集.docx"
    lookup = getattr(engine, "glossary", None).lookup
    if not root.is_dir() or root.is_symlink() or not glossary.is_file() or glossary.is_symlink():
        raise ValueError("source root invalid")
    if lookup("蒼樹会") != [("蒼樹会", ["医療法人社団 蒼樹会 みなみ野女性医療センター"])]:
        raise ValueError("project glossary binding changed")
    projects = [path for path in (root / "プロジェクト").iterdir() if path.is_dir() and not path.is_symlink() and _compact(path.name) == _compact("医療法人社団 蒼樹会 みなみ野女性医療センター")]
    if len(projects) != 1:
        raise ValueError("project not unique")
    project = projects[0]
    proposals = [path for path in (project / "00.提案").glob("*.pptx") if path.is_file() and not path.is_symlink() and not path.name.startswith("~$") and unicodedata.normalize("NFC", path.name) == "提案書.pptx"]
    schedules = [path for path in (project / "02.計画").glob("*.xlsx") if path.is_file() and not path.is_symlink() and not path.name.startswith("~$") and unicodedata.normalize("NFC", path.name) == "スケジュール.xlsx"]
    if len(proposals) != 1 or len(schedules) != 1:
        raise ValueError("proposal or schedule not unique")
    return root, glossary, proposals[0], schedules[0]


def _role_person(path: Path, role: str) -> str:
    if not zipfile.is_zipfile(path) or path.stat().st_size > 64 * 1024 * 1024:
        raise ValueError("proposal invalid")
    matches = []
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if re.fullmatch(r"ppt/slides/slide\d+\.xml", name) is None:
                continue
            root = ET.fromstring(archive.read(name))
            slide_number = int(re.search(r"\d+", name).group())
            for shape in root.findall(".//p:sp", _NS):
                text = "".join(node.text or "" for node in shape.findall(".//a:t", _NS)).strip()
                match = re.fullmatch(re.escape(role) + r"([一-鿿]{2,4}[\s　]+[一-鿿]{2,4})要件整理、業務論点整理、報告資料整備", text)
                if match:
                    matches.append((slide_number, unicodedata.normalize("NFKC", match.group(1)).strip()))
    if len(matches) != 1 or matches[0][0] != 8:
        raise ValueError("role-person roster edge not unique")
    return matches[0][1]


def _matching_task_ids(rows: Sequence[Sequence[object]], milestone: str, person: str) -> tuple[str, ...]:
    expected = ("タスクID", "担当者", "関連マイルストーン")
    if not rows:
        raise ValueError("schedule empty")
    header = tuple(str(value).strip() if value is not None else "" for value in rows[0])
    if any(header.count(label) != 1 for label in expected):
        raise ValueError("schedule headers ambiguous")
    task_col, person_col, milestone_col = (header.index(label) for label in expected)
    all_ids = set()
    matched = []
    for row in rows[1:]:
        if max(task_col, person_col, milestone_col) >= len(row):
            raise ValueError("schedule row truncated")
        task_id = unicodedata.normalize("NFKC", str(row[task_col] or "")).strip().upper()
        if re.fullmatch(r"T\d{2}", task_id) is None or task_id in all_ids:
            raise ValueError("task ID invalid or duplicate")
        all_ids.add(task_id)
        milestones = {token.strip().upper() for token in re.split(r"[,、/]", unicodedata.normalize("NFKC", str(row[milestone_col] or ""))) if token.strip()}
        people = {_compact(token) for token in re.split(r"[,、/]", unicodedata.normalize("NFKC", str(row[person_col] or ""))) if token.strip()}
        if milestone.upper() in milestones and _compact(person) in people:
            matched.append(task_id)
    if not matched:
        raise ValueError("no matching task")
    return tuple(matched)


def decide_question(engine: Any, question: str) -> StructuredCandidateDecision | None:
    contract = graph_contract_for_question(question)
    if contract is None:
        return None
    try:
        from openpyxl import load_workbook

        root, glossary, proposal, schedule = _sources(engine)
        before = schedule.read_bytes()
        person = _role_person(proposal, contract["bindings"]["role"])
        workbook = load_workbook(schedule, read_only=True, data_only=False)
        try:
            if workbook.sheetnames != ["スケジュール管理表"]:
                raise ValueError("schedule sheet set changed")
            rows = list(workbook["スケジュール管理表"].iter_rows(values_only=True))
        finally:
            workbook.close()
        if before != schedule.read_bytes():
            raise ValueError("schedule changed during read")
        task_ids = _matching_task_ids(rows, contract["bindings"]["milestone_id"], person)
        paths, digest = _fingerprint((glossary, proposal, schedule), root)
        result = StructuredCandidateAnswer("、".join(task_ids), paths, digest, len(contract["operation_graph"]["nodes"]), len(task_ids))
        return StructuredCandidateDecision("resolved", "certified_milestone_role_task_ids", result)
    except (ET.ParseError, ImportError, OSError, RuntimeError, TypeError, UnicodeError, ValueError, zipfile.BadZipFile):
        return StructuredCandidateDecision("hold", "milestone_role_task_not_certified")


__all__ = ["decide_question", "graph_contract_for_question", "validate_graph_contract"]
