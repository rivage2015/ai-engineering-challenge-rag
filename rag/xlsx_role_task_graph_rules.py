"""Fail-closed graph joins between workbook role and task-table evidence."""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from pathlib import Path
from typing import Any, Mapping

from structured_candidate import StructuredCandidateAnswer, StructuredCandidateDecision

VERSION = "0.1"
ROLE_TASK_COUNT = re.compile(
    r"^(?P<location>KSS)において、(?P<role>データエンジニア)"
    r"が担当するタスクIDはいくつありますか。$"
)


def _canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)


def _compact(value: object) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKC", str(value)).casefold()
        if not char.isspace()
    )


def graph_contract_for_question(question: str) -> dict[str, Any] | None:
    if not isinstance(question, str):
        return None
    match = ROLE_TASK_COUNT.fullmatch(question)
    if match is None:
        return None
    operators = (
        "bind_unique_project",
        "bind_current_proposal",
        "extract_role_person_edge_from_presentation",
        "bind_unique_project_schedule",
        "bind_resource_plan_sheet",
        "resolve_person_in_resource_plan",
        "extract_related_task_ids_from_role_row",
        "bind_wbs_task_sheet",
        "verify_every_related_task_id_exists",
        "verify_person_is_assigned_to_every_task_row",
        "deduplicate_task_ids",
        "count_task_ids",
    )
    nodes = []
    previous = "input_question"
    for index, operator in enumerate(operators, 1):
        output = f"value_{index:03d}"
        nodes.append(
            {
                "operation_id": f"op_{index:03d}_{operator}",
                "operator": operator,
                "input_refs": [previous],
                "output_ref": output,
            }
        )
        previous = output
    core = {
        "xlsx_role_task_graph_version": VERSION,
        "rule_id": "role_to_task_table_distinct_count",
        "question_sha256": hashlib.sha256(question.encode("utf-8")).hexdigest(),
        "bindings": match.groupdict(),
        "scope": {
            "source_channel": "xlsx_cell_values_and_cross_sheet_edges",
            "question_independent": True,
            "ambiguity_policy": "hold",
        },
        "operation_graph": {
            "external_inputs": [
                {
                    "input_ref": "input_question",
                    "input_type": "xlsx_workbook",
                    "source": "question_scope",
                }
            ],
            "nodes": nodes,
            "edges": [
                {"from": nodes[index - 1]["output_ref"], "to": nodes[index]["operation_id"]}
                for index in range(1, len(nodes))
            ],
        },
        "requested_output": {
            "source_operation_ref": nodes[-1]["operation_id"],
            "cardinality": "single",
            "answer_shape": {
                "container": "scalar",
                "value_type": "integer",
                "unit": None,
            },
            "display_precision": 0,
            "required_keys": None,
        },
    }
    return {
        "graph_contract_id": "xlsx_role_task_"
        + hashlib.sha256(_canonical(core).encode()).hexdigest()[:32],
        **core,
    }


def validate_graph_contract(question: str, contract: Mapping[str, Any]) -> bool:
    expected = graph_contract_for_question(question)
    return expected is not None and isinstance(contract, Mapping) and _canonical(expected) == _canonical(contract)


def _hold(reason: str) -> StructuredCandidateDecision:
    return StructuredCandidateDecision("hold", reason)


def decide_question(engine: Any, question: str) -> StructuredCandidateDecision | None:
    contract = graph_contract_for_question(question)
    if contract is None:
        return None
    try:
        from openpyxl import load_workbook
        from structured_candidate import _candidate_values, _location_matches

        root = Path(engine.source_root).resolve()
        if not root.is_dir() or root.is_symlink():
            return _hold("xlsx_role_task_source_root_invalid")
        candidates = _candidate_values(contract["bindings"]["location"], getattr(engine, "glossary", None))
        project_paths = []
        paths = []
        for path in root.rglob("*.xlsx"):
            if path.is_symlink() or not path.is_file() or path.name.startswith("~$"):
                continue
            relative = path.resolve().relative_to(root)
            if (
                unicodedata.normalize("NFC", path.name) == "スケジュール.xlsx"
                and _location_matches(relative.parts[:-1], candidates)
            ):
                paths.append(path)
        if len(paths) != 1:
            return _hold("xlsx_role_task_source_not_unique")
        path = paths[0]
        project = path.parents[1]
        proposals = [
            candidate
            for candidate in (project / "00.提案").iterdir()
            if candidate.is_file()
            and not candidate.is_symlink()
            and not candidate.name.startswith("~$")
            and unicodedata.normalize("NFC", candidate.name) == "提案書_final.pptx"
        ]
        if len(proposals) != 1:
            return _hold("xlsx_role_task_proposal_not_unique")
        proposal = proposals[0]
        from cross_document_finance_rules import _opc_text

        proposal_lines = [line.strip() for line in _opc_text(proposal).splitlines() if line.strip()]
        role_label = contract["bindings"]["role"]
        role_indexes = [index for index, line in enumerate(proposal_lines) if line == role_label]
        if len(role_indexes) != 1 or role_indexes[0] + 1 >= len(proposal_lines):
            return _hold("xlsx_role_task_proposal_role_not_unique")
        person = proposal_lines[role_indexes[0] + 1]
        if re.fullmatch(r"[\u4e00-\u9fff]{2,4}[\s　]+[\u4e00-\u9fff]{2,4}", person) is None:
            return _hold("xlsx_role_task_proposal_person_invalid")
        before = path.read_bytes()
        proposal_bytes = proposal.read_bytes()
        workbook = load_workbook(path, read_only=True, data_only=False)
        if set(workbook.sheetnames) != {
            "WBSタスク一覧",
            "マイルストーン",
            "会議体・チェックポイント",
            "リソース計画",
        }:
            return _hold("xlsx_role_task_sheet_set_invalid")
        resources = list(workbook["リソース計画"].iter_rows(values_only=True))
        wbs = list(workbook["WBSタスク一覧"].iter_rows(values_only=True))
        workbook.close()
        after = path.read_bytes()
        if before != after:
            return _hold("xlsx_role_task_source_changed")

        resource_header_rows = [index for index, row in enumerate(resources) if tuple(row[:3]) == ("担当者", "主担当領域", "関連タスク")]
        if resource_header_rows != [1]:
            return _hold("xlsx_role_task_resource_header_not_unique")
        person_rows = []
        for row in resources[resource_header_rows[0] + 1 :]:
            if not row[0] or not row[1] or not row[2]:
                continue
            if _compact(row[0]) == _compact(person):
                person_rows.append(row)
        if len(person_rows) != 1:
            return _hold("xlsx_role_task_person_not_unique")
        task_ids = tuple(re.findall(r"T[0-9]{2}", str(person_rows[0][2])))
        if not task_ids or len(task_ids) != len(set(task_ids)):
            return _hold("xlsx_role_task_related_ids_invalid")

        if not wbs or tuple(wbs[0][:6]) != (
            "タスクID",
            "フェーズNo.",
            "フェーズ名",
            "タスク名",
            "詳細・補足",
            "担当者",
        ):
            return _hold("xlsx_role_task_wbs_header_invalid")
        indexed = {}
        for row in wbs[1:]:
            task_id = str(row[0]).strip() if row[0] is not None else ""
            if not re.fullmatch(r"T[0-9]{2}", task_id) or task_id in indexed:
                return _hold("xlsx_role_task_wbs_id_invalid")
            indexed[task_id] = str(row[5] or "")
        if set(task_ids) - set(indexed):
            return _hold("xlsx_role_task_related_id_missing")
        if any(person not in indexed[task_id] for task_id in task_ids):
            return _hold("xlsx_role_task_assignment_edge_mismatch")

        relative_paths = tuple(
            unicodedata.normalize("NFC", source.relative_to(root).as_posix())
            for source in (proposal, path)
        )
        fingerprint = hashlib.sha256()
        for source, data in ((proposal, proposal_bytes), (path, before)):
            fingerprint.update(
                _canonical(
                    {
                        "relative_path": unicodedata.normalize("NFC", source.relative_to(root).as_posix()),
                        "sha256": hashlib.sha256(data).hexdigest(),
                        "size_bytes": len(data),
                    }
                ).encode()
            )
        result = StructuredCandidateAnswer(
            answer=str(len(task_ids)),
            source_paths=relative_paths,
            source_sha256=fingerprint.hexdigest(),
            operation_count=len(contract["operation_graph"]["nodes"]),
            output_count=1,
        )
        return StructuredCandidateDecision("resolved", "certified_xlsx_role_task_graph", result)
    except (AttributeError, OSError, RuntimeError, TypeError, ValueError):
        return _hold("xlsx_role_task_source_not_certified")


__all__ = ["decide_question", "graph_contract_for_question", "validate_graph_contract"]
