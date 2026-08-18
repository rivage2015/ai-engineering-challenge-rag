"""Fail-closed XLSX formula-lineage and linear-model metric rules.

The rules in this module intentionally answer only complete question grammars.
They never infer an answer from a filename, a question id, or an authored
cached result alone:

* the highlighted-formula lane follows a unique direct-yellow squared-residual
  formula through an INDEX/MATCH projection and a dated IF transform, then
  requires the recovered source value to agree with a unique raw record; and
* the model-metric lane aligns a unique labelled coefficient table with a
  population-standardised feature table, verifies every cached transform
  against its referenced source column and a raw table, scores every row, and
  maximises F1 over complete equal-score threshold groups.

Ambiguity, unsupported formulas, stale cached values, missing raw evidence, or
an uncertified live GraphPlan produces a hold rather than a guessed answer.
"""

from __future__ import annotations

import hashlib
import json
import math
import posixpath
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, localcontext
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Mapping, Sequence

from structured_candidate import (
    StructuredCandidateAnswer,
    StructuredCandidateDecision,
    _candidate_values,
    _location_matches,
)


XLSX_FORMULA_ML_RULE_VERSION = "0.1"

YELLOW_FORMULA_ATTRIBUTE = re.compile(
    r"^(?P<location>.+?)の(?P<container>[^,、。]+?\.xlsx)において、"
    r"(?P<color>[^,、。]+?)ハイライトセルは予測と実際の誤差を計算していますが、"
    r"その予測値の対象となっている(?P<entity>[^,、。]+?)の"
    r"(?P<attribute>建設年|建築年|竣工年)を算出してください。?$"
)

REGRESSION_F1_MAX = re.compile(
    r"^(?P<location>.+?)の(?P<container>[^,、。]+?)にて算出された回帰係数を用いて"
    r"全データの予測値を計算し、正解データに対する\s*F1\s*スコアが最大となるように"
    r"閾値を設定したときの\s*F1\s*スコアを答えてください。"
    r"小数第(?P<digits>[0-9０-９]+)位まで求めてください。?$"
)

_MAX_WORKBOOK_BYTES = 256 * 1024 * 1024
_MAX_ZIP_ENTRIES = 10_000
_MAX_MEMBER_BYTES = 128 * 1024 * 1024
_MAX_TOTAL_BYTES = 512 * 1024 * 1024
_MAX_COMPRESSION_RATIO = 500.0
_MAX_LINEAGE_CELLS = 4_000_000
_MAX_MODEL_CELLS = 500_000
_MAX_MODEL_ROWS = 200_000
_MAX_MODEL_COLUMNS = 512
_XML_FORBIDDEN = (b"<!DOCTYPE", b"<!ENTITY")
_YELLOW_NAMES = frozenset({"黄", "黄色", "yellow"})
_YELLOW_ARGB = "FFFFFF00"
_ARCHIVE_ENGLISH = re.compile(
    r"(?:^|[._\-\s])(?:old|draft|copy|backup|bak|archive|archived|obsolete|tmp)"
    r"(?:$|[._\-\s])",
    flags=re.IGNORECASE,
)
_ARCHIVE_JAPANESE = (
    "旧",
    "過去",
    "草案",
    "ドラフト",
    "コピー",
    "バックアップ",
    "アーカイブ",
)
_INTERCEPT_KEYS = frozenset({"切片", "intercept", "constant", "const"})
_OBSERVATION_KEYS = frozenset({"観測数", "observations", "observation count"})
_IDENTIFIER_KEYS = frozenset({"id", "record id", "record_id", "identifier", "識別子"})
_BUILT_YEAR_KEYS = frozenset(
    {
        "year built",
        "year_built",
        "建設年",
        "建築年",
        "竣工年",
    }
)
_BUILT_YEAR_DERIVED_KEYS = frozenset(
    {
        "year built fillna",
        "year built_fillna",
        "year_built fillna",
        "year_built_fillna",
        "建設年 fillna",
        "建設年_fillna",
    }
)
_REAL_ESTATE_ENTITY_KEYS = frozenset(
    {
        "不動産",
        "物件",
        "建物",
        "建築物",
        "家屋",
        "property",
        "real estate",
        "building",
    }
)
_REF_TOKEN = re.compile(
    r"(?:(?:'(?P<quoted>(?:[^']|'')+)'|(?P<plain>[^\s'!(),:+\-*/=<>]+))!)?"
    r"\$?(?P<column>[A-Z]{1,3})\$?(?P<row>[1-9][0-9]{0,6})",
    flags=re.IGNORECASE,
)
_LOCAL_CELL = re.compile(r"\$?([A-Z]{1,3})\$?([1-9][0-9]{0,6})", re.IGNORECASE)


class _InvalidSource(ValueError):
    pass


@dataclass(frozen=True)
class _Reference:
    sheet: str | None
    column: int
    row: int
    text: str


@dataclass(frozen=True)
class _SheetSnapshot:
    name: str
    state: str
    formula_rows: tuple[tuple[object, ...], ...]
    value_rows: tuple[tuple[object, ...], ...]


@dataclass(frozen=True)
class _CoefficientTable:
    sheet: str
    intercept: Decimal
    coefficients: Mapping[str, Decimal]
    labels: Mapping[str, str]
    observations: int


@dataclass(frozen=True)
class _StandardizedTable:
    sheet: str
    source_sheet: str
    feature_columns: Mapping[str, int]
    target_column: int
    target_label: str
    standardized: tuple[Mapping[str, Decimal], ...]
    targets: tuple[int, ...]


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )


def _normalized(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value)).casefold().strip()


def _field_key(value: object) -> str:
    rendered = re.sub(r"\s+", " ", _normalized(value))
    return rendered.replace("-", " ")


def _semantic_field_key(value: object) -> str:
    return re.sub(r"[\s_\-]+", " ", _normalized(value)).strip()


def _column_number(letters: str) -> int:
    result = 0
    for char in letters.upper():
        result = result * 26 + ord(char) - 64
    if not 1 <= result <= 16_384:
        raise _InvalidSource("xlsx_formula_ml_cell_reference_invalid")
    return result


def _column_letters(number: int) -> str:
    if not 1 <= number <= 16_384:
        raise _InvalidSource("xlsx_formula_ml_cell_reference_invalid")
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _decimal(value: object) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        result = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError, ValueError):
        return None
    return result if result.is_finite() else None


def _render_decimal(value: Decimal) -> str:
    rendered = format(value, "f")
    if "." in rendered:
        rendered = rendered.rstrip("0").rstrip(".")
    return rendered or "0"


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.lstrip().startswith("=")


def _is_archived_component(value: str) -> bool:
    rendered = _normalized(value)
    return bool(_ARCHIVE_ENGLISH.search(rendered)) or any(
        marker in rendered for marker in _ARCHIVE_JAPANESE
    )


def _source_root(engine: Any) -> Path | None:
    try:
        root = Path(engine.source_root)
        if not root.is_dir() or root.is_symlink():
            return None
        return root.resolve()
    except (AttributeError, OSError, RuntimeError, TypeError):
        return None


def _has_symlink_component(path: Path, root: Path) -> bool:
    current = path
    while True:
        if current.is_symlink():
            return True
        if current == root:
            return False
        if root not in current.parents:
            return True
        current = current.parent


def _named_workbooks(engine: Any, location: str, container: str) -> tuple[Path, ...]:
    root = _source_root(engine)
    if root is None:
        return ()
    locations = _candidate_values(location, getattr(engine, "glossary", None))
    names = {
        _normalized(candidate)
        for candidate in _candidate_values(container, getattr(engine, "glossary", None))
    }
    matches: list[Path] = []
    try:
        for path in root.rglob("*.xlsx"):
            if (
                not path.is_file()
                or path.is_symlink()
                or path.name.startswith(("~$", "."))
                or _has_symlink_component(path, root)
                or _normalized(path.name) not in names
            ):
                continue
            relative = path.relative_to(root)
            if any(_is_archived_component(part) for part in relative.parts):
                continue
            if not _location_matches(relative.parts[:-1], locations):
                continue
            size = path.stat().st_size
            if 0 < size <= _MAX_WORKBOOK_BYTES:
                matches.append(path.resolve())
    except OSError:
        return ()
    return tuple(sorted(set(matches), key=lambda item: item.as_posix()))


def _validate_archive(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        if not 1 <= len(infos) <= _MAX_ZIP_ENTRIES:
            raise _InvalidSource("xlsx_formula_ml_archive_invalid")
        seen: set[str] = set()
        total = 0
        for info in infos:
            name = info.filename
            pure = PurePosixPath(name)
            if (
                not name
                or "\\" in name
                or pure.is_absolute()
                or any(part in {"", ".", ".."} for part in pure.parts)
                or name in seen
                or info.flag_bits & 0x1
                or info.is_dir()
            ):
                raise _InvalidSource("xlsx_formula_ml_archive_invalid")
            seen.add(name)
            if not 0 <= info.file_size <= _MAX_MEMBER_BYTES:
                raise _InvalidSource("xlsx_formula_ml_archive_resource_limit")
            if info.file_size and info.compress_size == 0:
                raise _InvalidSource("xlsx_formula_ml_archive_invalid")
            if (
                info.compress_size
                and info.file_size / info.compress_size > _MAX_COMPRESSION_RATIO
            ):
                raise _InvalidSource("xlsx_formula_ml_archive_resource_limit")
            total += info.file_size
            if total > _MAX_TOTAL_BYTES:
                raise _InvalidSource("xlsx_formula_ml_archive_resource_limit")
            if name.casefold().endswith((".xml", ".rels")):
                data = archive.read(info)
                upper = data.upper()
                if any(marker in upper for marker in _XML_FORBIDDEN):
                    raise _InvalidSource("xlsx_formula_ml_xml_unsafe")
        required = {
            "[Content_Types].xml",
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
            "xl/styles.xml",
        }
        if not required.issubset(seen):
            raise _InvalidSource("xlsx_formula_ml_archive_invalid")


def _nodes(operators: Sequence[str]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    previous = "input_question"
    for index, operator in enumerate(operators, 1):
        output = f"value_{index:03d}"
        result.append(
            {
                "operation_id": f"op_{index:03d}_{operator}",
                "operator": operator,
                "input_refs": [previous],
                "output_ref": output,
            }
        )
        previous = output
    return result


def _contract(
    question: str,
    rule_id: str,
    bindings: Mapping[str, Any],
    scope: Mapping[str, Any],
    operators: Sequence[str],
    value_type: str,
    display_precision: Mapping[str, Any] | None,
) -> dict[str, Any]:
    nodes = _nodes(operators)
    core = {
        "graph_rule_version": XLSX_FORMULA_ML_RULE_VERSION,
        "rule_id": rule_id,
        "question_sha256": hashlib.sha256(question.encode("utf-8")).hexdigest(),
        "bindings": dict(bindings),
        "scope": dict(scope),
        "operation_graph": {
            "external_inputs": [
                {
                    "input_ref": "input_question",
                    "input_type": "source_records",
                    "source": "question_scope",
                }
            ],
            "nodes": nodes,
            "edges": [
                {"from": nodes[index - 1]["output_ref"], "to": nodes[index]["operation_id"]}
                for index in range(1, len(nodes))
            ],
        },
        "requested_output": {
            "source_operation_ref": nodes[-1]["operation_id"],
            "cardinality": "single",
            "answer_shape": {
                "container": "scalar",
                "value_type": value_type,
                "unit": None,
            },
            "display_precision": dict(display_precision) if display_precision else None,
            "required_keys": None,
        },
    }
    return {
        "graph_contract_id": "xlsx_formula_ml_"
        + hashlib.sha256(_canonical_json(core).encode("utf-8")).hexdigest()[:32],
        **core,
    }


def graph_contract_for_question(question: str) -> dict[str, Any] | None:
    if not isinstance(question, str):
        return None
    lineage = YELLOW_FORMULA_ATTRIBUTE.fullmatch(question)
    if lineage is not None:
        if (
            _normalized(lineage["color"]) not in _YELLOW_NAMES
            or _semantic_field_key(lineage["entity"])
            not in {
                _semantic_field_key(value)
                for value in _REAL_ESTATE_ENTITY_KEYS
            }
        ):
            return None
        bindings = {
            key: lineage[key]
            for key in ("location", "container", "color", "entity", "attribute")
        }
        return _contract(
            question,
            "xlsx_yellow_residual_formula_lineage_attribute",
            bindings,
            {
                "location": bindings["location"],
                "container": bindings["container"],
                "style_channel": "direct_solid_fill",
                "color": bindings["color"],
                "formula_semantics": "squared_prediction_residual",
                "lineage_verification": "derived_transform_and_unique_raw_record",
            },
            (
                "retrieve",
                "locate_unique_highlight",
                "parse_formula",
                "resolve_precedents",
                "invert_transform",
                "select_unique_raw_record",
                "verify_exact",
                "project",
            ),
            "integer",
            None,
        )
    metric = REGRESSION_F1_MAX.fullmatch(question)
    if metric is None:
        return None
    digits = int(unicodedata.normalize("NFKC", metric["digits"]))
    if not 0 <= digits <= 12:
        return None
    bindings = {
        "location": metric["location"],
        "container": metric["container"],
        "digits": metric["digits"],
    }
    return _contract(
        question,
        "xlsx_standardized_linear_prediction_f1_argmax",
        bindings,
        {
            "location": bindings["location"],
            "container": bindings["container"],
            "coefficient_source": "unique_labeled_coefficient_table",
            "feature_transform": "STANDARDIZE_with_population_stdev",
            "threshold_domain": "complete_unique_prediction_groups",
            "threshold_tie_break": "highest_threshold",
            "metric": "binary_f1",
        },
        (
            "retrieve",
            "select_coefficient_table",
            "align_fields",
            "verify_standardization",
            "calculate_predictions",
            "group_equal_scores",
            "calculate_f1",
            "argmax_all",
            "project",
            "round",
        ),
        "number",
        {"mode": "decimal_places", "digits": digits},
    )


def validate_graph_contract(question: str, contract: Mapping[str, Any]) -> bool:
    expected = graph_contract_for_question(question)
    if expected is None or not isinstance(contract, Mapping):
        return False
    try:
        return _canonical_json(expected) == _canonical_json(contract)
    except (TypeError, ValueError):
        return False


def _resolve_sheet_name(workbook: Any, requested: str) -> str:
    matches = [
        name for name in workbook.sheetnames if _normalized(name) == _normalized(requested)
    ]
    if len(matches) != 1:
        raise _InvalidSource("xlsx_formula_ml_sheet_reference_ambiguous")
    return matches[0]


def _references(formula: str, local_sheet: str) -> tuple[_Reference, ...]:
    result: list[_Reference] = []
    for match in _REF_TOKEN.finditer(unicodedata.normalize("NFKC", formula)):
        sheet = match["quoted"] or match["plain"]
        if sheet is not None:
            sheet = sheet.replace("''", "'")
        result.append(
            _Reference(
                sheet=sheet or local_sheet,
                column=_column_number(match["column"]),
                row=int(match["row"]),
                text=match.group(0),
            )
        )
    return tuple(result)


def _sheet_rows(
    formula_workbook: Any,
    value_workbook: Any,
    sheet_name: str,
    requested_rows: Iterable[int],
) -> dict[int, tuple[tuple[object, object], ...]]:
    rows = frozenset(requested_rows)
    if not rows or min(rows) < 1:
        raise _InvalidSource("xlsx_formula_ml_cell_reference_invalid")
    formula_sheet = formula_workbook[sheet_name]
    value_sheet = value_workbook[sheet_name]
    result: dict[int, tuple[tuple[object, object], ...]] = {}
    for row_number, (formula_row, value_row) in enumerate(
        zip(formula_sheet.iter_rows(), value_sheet.iter_rows()), start=1
    ):
        if row_number in rows:
            width = max(len(formula_row), len(value_row))
            result[row_number] = tuple(
                (
                    formula_row[index].value if index < len(formula_row) else None,
                    value_row[index].value if index < len(value_row) else None,
                )
                for index in range(width)
            )
        if row_number >= max(rows):
            break
    if set(result) != set(rows):
        raise _InvalidSource("xlsx_formula_ml_cell_reference_invalid")
    return result


def _row_value(
    rows: Mapping[int, tuple[tuple[object, object], ...]],
    row: int,
    column: int,
    *,
    cached: bool,
) -> object:
    values = rows.get(row)
    if values is None or not 1 <= column <= len(values):
        return None
    return values[column - 1][1 if cached else 0]


def _yellow_cells(workbook: Any) -> list[tuple[str, str, str]]:
    total = 0
    result: list[tuple[str, str, str]] = []
    for sheet in workbook.worksheets:
        rows = int(sheet.max_row or 0)
        columns = int(sheet.max_column or 0)
        if rows < 0 or columns < 0 or rows > 1_048_576 or columns > 16_384:
            raise _InvalidSource("xlsx_formula_ml_worksheet_dimension_invalid")
        total += rows * columns
        if total > _MAX_LINEAGE_CELLS:
            raise _InvalidSource("xlsx_formula_ml_worksheet_resource_limit")
        if sheet.sheet_state != "visible":
            continue
        for row in sheet.iter_rows():
            for cell in row:
                fill = getattr(cell, "fill", None)
                foreground = getattr(fill, "fgColor", None)
                if (
                    getattr(fill, "fill_type", None) == "solid"
                    and getattr(foreground, "type", None) == "rgb"
                    and str(getattr(foreground, "rgb", "")).upper() == _YELLOW_ARGB
                ):
                    coordinate = getattr(cell, "coordinate", None)
                    if not isinstance(coordinate, str) or not _is_formula(cell.value):
                        raise _InvalidSource("xlsx_formula_ml_highlight_not_formula")
                    result.append((sheet.title, coordinate, str(cell.value)))
    return result


def _parse_residual_formula(
    formula: str, local_sheet: str
) -> tuple[str, _Reference, tuple[_Reference, ...]]:
    compact = re.sub(r"\s+", "", unicodedata.normalize("NFKC", formula))
    if compact.startswith("="):
        compact = compact[1:]
    reference_pattern = (
        r"(?:(?:'(?:[^']|'')+'|[^\s'!(),:+\-*/=<>]+)!)?"
        r"\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6}"
    )
    match = re.fullmatch(
        rf"\((?P<prediction>.+)-(?P<actual>{reference_pattern})\)\^2",
        compact,
        flags=re.IGNORECASE,
    )
    if match is None or "*" not in match["prediction"] or "+" not in match["prediction"]:
        raise _InvalidSource("xlsx_formula_ml_residual_formula_unsupported")
    actual_refs = _references(match["actual"], local_sheet)
    prediction_refs = _references(match["prediction"], local_sheet)
    if len(actual_refs) != 1 or len(prediction_refs) < 3:
        raise _InvalidSource("xlsx_formula_ml_residual_formula_unsupported")
    return match["prediction"], actual_refs[0], prediction_refs


def _attribute_matches_derived(header: object, attribute: str) -> bool:
    if _normalized(attribute) not in {"建設年", "建築年", "竣工年"}:
        return False
    return _semantic_field_key(header) in {
        _semantic_field_key(value) for value in _BUILT_YEAR_DERIVED_KEYS
    }


def _attribute_matches_raw(header: object, attribute: str) -> bool:
    if _normalized(attribute) not in {"建設年", "建築年", "竣工年"}:
        return False
    return _semantic_field_key(header) in {
        _semantic_field_key(value) for value in _BUILT_YEAR_KEYS
    }


def _parse_index_match(
    formula: object,
    row_number: int,
    header_row: Sequence[tuple[object, object]],
    data_row: Sequence[tuple[object, object]],
) -> int:
    if not _is_formula(formula):
        raise _InvalidSource("xlsx_formula_ml_projection_formula_unsupported")
    compact = re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(formula)))
    match = re.fullmatch(
        r"=INDEX\((?P<a>\$?[A-Z]{1,3}\$?[0-9]+):(?P<b>\$?[A-Z]{1,3}\$?[0-9]+),"
        r"1,MATCH\((?P<h>\$?[A-Z]{1,3}\$?1),"
        r"(?P<ha>\$?[A-Z]{1,3}\$?1):(?P<hb>\$?[A-Z]{1,3}\$?1),0\)\)",
        compact,
        flags=re.IGNORECASE,
    )
    if match is None:
        raise _InvalidSource("xlsx_formula_ml_projection_formula_unsupported")

    def cell(value: str) -> tuple[int, int]:
        parsed = _LOCAL_CELL.fullmatch(value)
        if parsed is None:
            raise _InvalidSource("xlsx_formula_ml_projection_formula_unsupported")
        return int(parsed.group(2)), _column_number(parsed.group(1))

    a_row, a_column = cell(match["a"])
    b_row, b_column = cell(match["b"])
    h_row, h_column = cell(match["h"])
    ha_row, ha_column = cell(match["ha"])
    hb_row, hb_column = cell(match["hb"])
    if (
        a_row != row_number
        or b_row != row_number
        or h_row != 1
        or ha_row != 1
        or hb_row != 1
        or a_column > b_column
        or ha_column != a_column
        or hb_column != b_column
        or h_column > len(header_row)
    ):
        raise _InvalidSource("xlsx_formula_ml_projection_formula_unsupported")
    lookup = header_row[h_column - 1][1]
    matches = [
        column
        for column in range(a_column, b_column + 1)
        if column <= len(header_row)
        and _field_key(header_row[column - 1][1]) == _field_key(lookup)
    ]
    if len(matches) != 1 or matches[0] > len(data_row):
        raise _InvalidSource("xlsx_formula_ml_projection_header_ambiguous")
    return matches[0]


def _parse_year_transform(
    formula: object,
    row_number: int,
    formula_row: Sequence[tuple[object, object]],
) -> tuple[Decimal, Decimal]:
    if not _is_formula(formula):
        raise _InvalidSource("xlsx_formula_ml_transform_formula_unsupported")
    compact = re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(formula)))
    refs = _references(compact, "")
    if len(refs) != 2 or refs[0].column != refs[1].column or refs[0].row != refs[1].row:
        raise _InvalidSource("xlsx_formula_ml_transform_formula_unsupported")
    canonical = _REF_TOKEN.sub("REF", compact).upper()
    match = re.fullmatch(
        r"=(?P<base>[+\-]?\d+(?:\.\d+)?)-IF\(REF>0,REF,(?P<fallback>[+\-]?\d+(?:\.\d+)?)\)",
        canonical,
    )
    source = refs[0]
    if match is None or source.row != row_number or source.column > len(formula_row):
        raise _InvalidSource("xlsx_formula_ml_transform_formula_unsupported")
    base = _decimal(match["base"])
    fallback = _decimal(match["fallback"])
    raw = _decimal(formula_row[source.column - 1][1])
    if base is None or fallback is None or raw is None:
        raise _InvalidSource("xlsx_formula_ml_transform_value_invalid")
    expected = base - (raw if raw > 0 else fallback)
    return raw, expected


def _unique_identifier_column(headers: Sequence[tuple[object, object]]) -> int:
    matches = [
        index
        for index, pair in enumerate(headers, start=1)
        if _semantic_field_key(pair[1]) in {
            _semantic_field_key(value) for value in _IDENTIFIER_KEYS
        }
    ]
    if len(matches) != 1:
        raise _InvalidSource("xlsx_formula_ml_identifier_ambiguous")
    return matches[0]


def _verify_unique_raw_record(
    value_workbook: Any,
    derived_sheet: str,
    identifier_header: object,
    identifier_value: object,
    attribute: str,
    expected: Decimal,
) -> None:
    matching_tables: list[tuple[str, int, int]] = []
    for sheet in value_workbook.worksheets:
        if sheet.sheet_state != "visible" or sheet.title == derived_sheet:
            continue
        first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first is None:
            continue
        identifier_columns = [
            index
            for index, value in enumerate(first)
            if _field_key(value) == _field_key(identifier_header)
        ]
        attribute_columns = [
            index
            for index, value in enumerate(first)
            if _attribute_matches_raw(value, attribute)
        ]
        if len(identifier_columns) == 1 and len(attribute_columns) == 1:
            matching_tables.append(
                (sheet.title, identifier_columns[0], attribute_columns[0])
            )
    if len(matching_tables) != 1:
        raise _InvalidSource("xlsx_formula_ml_raw_table_not_unique")
    sheet_name, identifier_column, attribute_column = matching_tables[0]
    matches: list[Decimal] = []
    for row in value_workbook[sheet_name].iter_rows(min_row=2, values_only=True):
        if identifier_column >= len(row) or attribute_column >= len(row):
            continue
        if str(row[identifier_column]) != str(identifier_value):
            continue
        value = _decimal(row[attribute_column])
        if value is None:
            raise _InvalidSource("xlsx_formula_ml_raw_value_invalid")
        matches.append(value)
    if len(matches) != 1 or matches[0] != expected:
        raise _InvalidSource("xlsx_formula_ml_raw_record_mismatch")


def _lineage_answer(path: Path, attribute: str) -> str:
    from openpyxl import load_workbook

    formula_workbook = load_workbook(
        path, read_only=True, data_only=False, keep_links=False
    )
    value_workbook = load_workbook(
        path, read_only=True, data_only=True, keep_links=False
    )
    try:
        yellow = _yellow_cells(formula_workbook)
        if len(yellow) != 1:
            raise _InvalidSource("xlsx_formula_ml_highlight_not_unique")
        formula_sheet, _, formula = yellow[0]
        _, actual, prediction_refs = _parse_residual_formula(formula, formula_sheet)
        actual_sheet = _resolve_sheet_name(formula_workbook, actual.sheet or formula_sheet)
        actual = _Reference(actual_sheet, actual.column, actual.row, actual.text)
        data_refs: list[_Reference] = []
        for reference in prediction_refs:
            sheet_name = _resolve_sheet_name(
                formula_workbook, reference.sheet or formula_sheet
            )
            if (
                sheet_name == actual.sheet
                and reference.row == actual.row
            ):
                data_refs.append(
                    _Reference(sheet_name, reference.column, reference.row, reference.text)
                )
        if not data_refs:
            raise _InvalidSource("xlsx_formula_ml_prediction_row_not_found")
        rows = _sheet_rows(
            formula_workbook,
            value_workbook,
            actual.sheet or "",
            (1, actual.row),
        )
        header_row = rows[1]
        data_row = rows[actual.row]
        if actual.column > len(data_row) or _decimal(data_row[actual.column - 1][1]) is None:
            raise _InvalidSource("xlsx_formula_ml_actual_value_invalid")
        candidates = [
            reference
            for reference in data_refs
            if reference.column <= len(header_row)
            and _attribute_matches_derived(
                header_row[reference.column - 1][1], attribute
            )
        ]
        if len(candidates) != 1:
            raise _InvalidSource("xlsx_formula_ml_attribute_precedent_not_unique")
        projected = candidates[0]
        projected_formula = _row_value(
            rows, projected.row, projected.column, cached=False
        )
        projected_cached = _decimal(
            _row_value(rows, projected.row, projected.column, cached=True)
        )
        if projected_cached is None:
            raise _InvalidSource("xlsx_formula_ml_projection_cache_missing")
        source_column = _parse_index_match(
            projected_formula,
            projected.row,
            header_row,
            data_row,
        )
        source_header = header_row[source_column - 1][1]
        source_formula = data_row[source_column - 1][0]
        source_cached = _decimal(data_row[source_column - 1][1])
        if (
            not _attribute_matches_derived(source_header, attribute)
            or source_cached is None
            or source_cached != projected_cached
        ):
            raise _InvalidSource("xlsx_formula_ml_projection_cache_mismatch")
        raw_value, expected_transform = _parse_year_transform(
            source_formula, projected.row, data_row
        )
        if expected_transform != source_cached:
            raise _InvalidSource("xlsx_formula_ml_transform_cache_mismatch")
        raw_candidates = [
            index
            for index, pair in enumerate(header_row, start=1)
            if _attribute_matches_raw(pair[1], attribute)
        ]
        if len(raw_candidates) != 1:
            raise _InvalidSource("xlsx_formula_ml_raw_attribute_ambiguous")
        raw_column = raw_candidates[0]
        raw_direct = _decimal(data_row[raw_column - 1][1])
        if raw_direct is None or raw_direct != raw_value:
            raise _InvalidSource("xlsx_formula_ml_transform_source_mismatch")
        identifier_column = _unique_identifier_column(header_row)
        identifier_header = header_row[identifier_column - 1][1]
        identifier_value = data_row[identifier_column - 1][1]
        if identifier_value is None or str(identifier_value).strip() == "":
            raise _InvalidSource("xlsx_formula_ml_identifier_missing")
        _verify_unique_raw_record(
            value_workbook,
            actual.sheet or "",
            identifier_header,
            identifier_value,
            attribute,
            raw_value,
        )
    finally:
        formula_workbook.close()
        value_workbook.close()
    if raw_value != raw_value.to_integral_value():
        raise _InvalidSource("xlsx_formula_ml_attribute_not_integer")
    return _render_decimal(raw_value)


def _snapshot_workbook(formula_workbook: Any, value_workbook: Any) -> tuple[_SheetSnapshot, ...]:
    if formula_workbook.sheetnames != value_workbook.sheetnames:
        raise _InvalidSource("xlsx_formula_ml_workbook_pair_mismatch")
    total = 0
    result: list[_SheetSnapshot] = []
    for formula_sheet, value_sheet in zip(
        formula_workbook.worksheets, value_workbook.worksheets
    ):
        rows = max(int(formula_sheet.max_row or 0), int(value_sheet.max_row or 0))
        columns = max(
            int(formula_sheet.max_column or 0), int(value_sheet.max_column or 0)
        )
        if (
            rows < 1
            or columns < 1
            or rows > _MAX_MODEL_ROWS
            or columns > _MAX_MODEL_COLUMNS
        ):
            raise _InvalidSource("xlsx_formula_ml_model_dimension_invalid")
        total += rows * columns
        if total > _MAX_MODEL_CELLS:
            raise _InvalidSource("xlsx_formula_ml_model_resource_limit")
        formula_rows: list[tuple[object, ...]] = []
        value_rows: list[tuple[object, ...]] = []
        formula_iterator = formula_sheet.iter_rows(values_only=True)
        value_iterator = value_sheet.iter_rows(values_only=True)
        for formula_row, value_row in zip(formula_iterator, value_iterator):
            width = max(len(formula_row), len(value_row), columns)
            formula_rows.append(
                tuple(
                    formula_row[index] if index < len(formula_row) else None
                    for index in range(width)
                )
            )
            value_rows.append(
                tuple(
                    value_row[index] if index < len(value_row) else None
                    for index in range(width)
                )
            )
        if len(formula_rows) != rows or len(value_rows) != rows:
            raise _InvalidSource("xlsx_formula_ml_model_row_mismatch")
        result.append(
            _SheetSnapshot(
                name=formula_sheet.title,
                state=formula_sheet.sheet_state,
                formula_rows=tuple(formula_rows),
                value_rows=tuple(value_rows),
            )
        )
    return tuple(result)


def _coefficient_table(sheets: Sequence[_SheetSnapshot]) -> _CoefficientTable:
    headers: list[tuple[_SheetSnapshot, int, int]] = []
    for sheet in sheets:
        if sheet.state != "visible":
            continue
        for row_index, row in enumerate(sheet.formula_rows):
            for column_index, value in enumerate(row):
                if _normalized(value) in {"係数", "coefficient", "coefficients"}:
                    headers.append((sheet, row_index, column_index))
    if len(headers) != 1:
        raise _InvalidSource("xlsx_formula_ml_coefficient_table_not_unique")
    sheet, header_row, coefficient_column = headers[0]
    if coefficient_column < 1:
        raise _InvalidSource("xlsx_formula_ml_coefficient_table_invalid")
    label_column = coefficient_column - 1
    intercept: Decimal | None = None
    coefficients: dict[str, Decimal] = {}
    labels: dict[str, str] = {}
    for row_index in range(header_row + 1, len(sheet.formula_rows)):
        label = sheet.formula_rows[row_index][label_column]
        formula_value = sheet.formula_rows[row_index][coefficient_column]
        cached_value = sheet.value_rows[row_index][coefficient_column]
        if label is None and formula_value is None and cached_value is None:
            break
        if not isinstance(label, str) or not label.strip() or _is_formula(formula_value):
            raise _InvalidSource("xlsx_formula_ml_coefficient_table_invalid")
        number = _decimal(cached_value)
        if number is None:
            raise _InvalidSource("xlsx_formula_ml_coefficient_table_invalid")
        key = _semantic_field_key(label)
        if key in {_semantic_field_key(value) for value in _INTERCEPT_KEYS}:
            if intercept is not None:
                raise _InvalidSource("xlsx_formula_ml_intercept_not_unique")
            intercept = number
        else:
            if not key or key in coefficients:
                raise _InvalidSource("xlsx_formula_ml_coefficient_field_ambiguous")
            coefficients[key] = number
            labels[key] = label.strip()
    if intercept is None or not coefficients:
        raise _InvalidSource("xlsx_formula_ml_coefficient_table_invalid")
    observations: list[int] = []
    for row_index, row in enumerate(sheet.formula_rows):
        for column_index, value in enumerate(row[:-1]):
            if _semantic_field_key(value) not in {
                _semantic_field_key(item) for item in _OBSERVATION_KEYS
            }:
                continue
            number = _decimal(sheet.value_rows[row_index][column_index + 1])
            if number is not None and number == number.to_integral_value() and number > 0:
                observations.append(int(number))
    if len(observations) != 1:
        raise _InvalidSource("xlsx_formula_ml_observation_count_not_unique")
    return _CoefficientTable(
        sheet=sheet.name,
        intercept=intercept,
        coefficients=coefficients,
        labels=labels,
        observations=observations[0],
    )


def _parse_standardize_formula(
    formula: object,
    current_row: int,
    expected_last_row: int,
) -> tuple[str, int]:
    if not _is_formula(formula):
        raise _InvalidSource("xlsx_formula_ml_standardize_formula_missing")
    compact = re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(formula)))
    canonical = _REF_TOKEN.sub("REF", compact).upper()
    if canonical not in {
        "=STANDARDIZE(REF,AVERAGE(REF:REF),_XLFN.STDEV.P(REF:REF))",
        "=STANDARDIZE(REF,AVERAGE(REF:REF),STDEV.P(REF:REF))",
    }:
        raise _InvalidSource("xlsx_formula_ml_standardize_formula_unsupported")
    refs = _references(compact, "")
    if (
        len(refs) != 5
        or not refs[0].sheet
        or not refs[1].sheet
        or not refs[3].sheet
    ):
        raise _InvalidSource("xlsx_formula_ml_standardize_formula_unsupported")
    # In an Excel range such as ``Source!A$2:A$99`` the qualifier on the
    # first endpoint applies to the unqualified second endpoint.
    effective_sheets = (
        refs[0].sheet,
        refs[1].sheet,
        refs[2].sheet or refs[1].sheet,
        refs[3].sheet,
        refs[4].sheet or refs[3].sheet,
    )
    sheets = {_normalized(sheet) for sheet in effective_sheets}
    columns = {reference.column for reference in refs}
    if (
        len(sheets) != 1
        or len(columns) != 1
        or refs[0].row != current_row
        or refs[1].row != 2
        or refs[2].row != expected_last_row
        or refs[3].row != 2
        or refs[4].row != expected_last_row
    ):
        raise _InvalidSource("xlsx_formula_ml_standardize_range_invalid")
    return str(refs[0].sheet), refs[0].column


def _standardized_table(
    sheets: Sequence[_SheetSnapshot], coefficient: _CoefficientTable
) -> _StandardizedTable:
    candidates: list[_StandardizedTable] = []
    by_name = {_normalized(sheet.name): sheet for sheet in sheets}
    if len(by_name) != len(sheets):
        raise _InvalidSource("xlsx_formula_ml_sheet_name_ambiguous")
    feature_keys = frozenset(coefficient.coefficients)
    for sheet in sheets:
        if sheet.state != "visible" or len(sheet.formula_rows) < 2:
            continue
        header = sheet.value_rows[0]
        populated = [index for index, value in enumerate(header) if value is not None and str(value).strip()]
        keys = [_semantic_field_key(header[index]) for index in populated]
        if len(keys) != len(set(keys)) or not feature_keys.issubset(keys):
            continue
        extras = [key for key in keys if key not in feature_keys]
        if len(populated) != len(feature_keys) + 1 or len(extras) != 1:
            continue
        feature_columns = {key: keys.index(key) for key in feature_keys}
        target_key = extras[0]
        target_column = keys.index(target_key)
        # keys indexes are within populated; require the authored table to be
        # contiguous so column arithmetic cannot silently select blank gaps.
        if populated != list(range(len(populated))):
            continue
        row_count = len(sheet.formula_rows) - 1
        if row_count != coefficient.observations:
            continue
        source_name: str | None = None
        source_columns: dict[str, int] = {}
        standardized: list[Mapping[str, Decimal]] = []
        targets: list[int] = []
        valid = True
        for row_offset in range(1, len(sheet.formula_rows)):
            row_number = row_offset + 1
            values: dict[str, Decimal] = {}
            for key, column in feature_columns.items():
                try:
                    formula_source, source_column = _parse_standardize_formula(
                        sheet.formula_rows[row_offset][column],
                        row_number,
                        len(sheet.formula_rows),
                    )
                except _InvalidSource:
                    valid = False
                    break
                if source_name is None:
                    source_name = formula_source
                if _normalized(source_name) != _normalized(formula_source):
                    valid = False
                    break
                if key in source_columns and source_columns[key] != source_column:
                    valid = False
                    break
                source_columns[key] = source_column
                number = _decimal(sheet.value_rows[row_offset][column])
                if number is None:
                    valid = False
                    break
                values[key] = number
            if not valid:
                break
            target_formula = sheet.formula_rows[row_offset][target_column]
            target = _decimal(sheet.value_rows[row_offset][target_column])
            if _is_formula(target_formula) or target not in {Decimal(0), Decimal(1)}:
                valid = False
                break
            standardized.append(values)
            targets.append(int(target))
        if not valid or source_name is None or not targets or len(set(targets)) != 2:
            continue
        source = by_name.get(_normalized(source_name))
        if source is None or source.state != "visible" or len(source.value_rows) != len(sheet.value_rows):
            continue
        source_header = source.value_rows[0]
        if any(
            column > len(source_header)
            or _semantic_field_key(source_header[column - 1]) != key
            for key, column in source_columns.items()
        ):
            continue
        source_target_matches = [
            index
            for index, value in enumerate(source_header)
            if _semantic_field_key(value) == target_key
        ]
        if len(source_target_matches) != 1:
            continue
        source_target_column = source_target_matches[0]
        raw_features: dict[str, list[float]] = {key: [] for key in feature_keys}
        source_valid = True
        for row_offset in range(1, len(source.value_rows)):
            for key, column in source_columns.items():
                if _is_formula(source.formula_rows[row_offset][column - 1]):
                    source_valid = False
                    break
                number = _decimal(source.value_rows[row_offset][column - 1])
                if number is None:
                    source_valid = False
                    break
                raw_features[key].append(float(number))
            if not source_valid:
                break
            target = _decimal(source.value_rows[row_offset][source_target_column])
            if target is None or int(target) != targets[row_offset - 1] or target not in {Decimal(0), Decimal(1)}:
                source_valid = False
                break
        if not source_valid:
            continue
        for key, raw_values in raw_features.items():
            mean = math.fsum(raw_values) / len(raw_values)
            variance = math.fsum((value - mean) ** 2 for value in raw_values) / len(raw_values)
            deviation = math.sqrt(variance)
            if not math.isfinite(deviation) or deviation <= 0:
                source_valid = False
                break
            for index, raw_value in enumerate(raw_values):
                expected = (raw_value - mean) / deviation
                observed = float(standardized[index][key])
                if not math.isclose(observed, expected, rel_tol=1e-10, abs_tol=1e-10):
                    source_valid = False
                    break
            if not source_valid:
                break
        if not source_valid:
            continue
        _verify_model_raw_table(
            sheets,
            excluded={coefficient.sheet, sheet.name, source.name},
            source=source,
            feature_keys=feature_keys,
            target_key=target_key,
            source_columns=source_columns,
            source_target_column=source_target_column,
        )
        candidates.append(
            _StandardizedTable(
                sheet=sheet.name,
                source_sheet=source.name,
                feature_columns=feature_columns,
                target_column=target_column,
                target_label=str(header[target_column]),
                standardized=tuple(standardized),
                targets=tuple(targets),
            )
        )
    if len(candidates) != 1:
        raise _InvalidSource("xlsx_formula_ml_standardized_table_not_unique")
    return candidates[0]


def _verify_model_raw_table(
    sheets: Sequence[_SheetSnapshot],
    *,
    excluded: set[str],
    source: _SheetSnapshot,
    feature_keys: frozenset[str],
    target_key: str,
    source_columns: Mapping[str, int],
    source_target_column: int,
) -> None:
    candidates: list[tuple[_SheetSnapshot, Mapping[str, int], int]] = []
    for sheet in sheets:
        if sheet.state != "visible" or sheet.name in excluded or not sheet.value_rows:
            continue
        headers = [_semantic_field_key(value) for value in sheet.value_rows[0]]
        if len([value for value in headers if value]) != len(set(value for value in headers if value)):
            continue
        if not feature_keys.issubset(headers) or target_key not in headers:
            continue
        candidates.append(
            (
                sheet,
                {key: headers.index(key) for key in feature_keys},
                headers.index(target_key),
            )
        )
    if len(candidates) != 1:
        raise _InvalidSource("xlsx_formula_ml_raw_model_table_not_unique")
    raw, raw_columns, raw_target = candidates[0]
    if len(raw.value_rows) != len(source.value_rows):
        raise _InvalidSource("xlsx_formula_ml_raw_model_row_mismatch")
    for row_index in range(1, len(source.value_rows)):
        for key in feature_keys:
            left = _decimal(source.value_rows[row_index][source_columns[key] - 1])
            right = _decimal(raw.value_rows[row_index][raw_columns[key]])
            if left is None or right is None or left != right:
                raise _InvalidSource("xlsx_formula_ml_raw_model_value_mismatch")
        left_target = _decimal(source.value_rows[row_index][source_target_column])
        right_target = _decimal(raw.value_rows[row_index][raw_target])
        if left_target is None or right_target is None or left_target != right_target:
            raise _InvalidSource("xlsx_formula_ml_raw_model_value_mismatch")


def _maximum_f1(
    coefficient: _CoefficientTable,
    standardized: _StandardizedTable,
) -> Decimal:
    scored: list[tuple[Decimal, int]] = []
    for values, target in zip(standardized.standardized, standardized.targets):
        prediction = coefficient.intercept
        for key, weight in coefficient.coefficients.items():
            if key not in values:
                raise _InvalidSource("xlsx_formula_ml_prediction_field_missing")
            prediction += weight * values[key]
        if not prediction.is_finite():
            raise _InvalidSource("xlsx_formula_ml_prediction_invalid")
        scored.append((prediction, target))
    positives = sum(target for _, target in scored)
    if positives <= 0 or positives >= len(scored):
        raise _InvalidSource("xlsx_formula_ml_binary_target_invalid")
    scored.sort(key=lambda item: item[0], reverse=True)
    best_numerator = -1
    best_denominator = 1
    true_positives = 0
    index = 0
    while index < len(scored):
        end = index + 1
        group_positives = scored[index][1]
        while end < len(scored) and scored[end][0] == scored[index][0]:
            group_positives += scored[end][1]
            end += 1
        true_positives += group_positives
        numerator = 2 * true_positives
        denominator = positives + end
        # Strict comparison deliberately preserves the first (highest)
        # threshold when distinct score groups have exactly equal F1.
        if best_numerator < 0 or numerator * best_denominator > best_numerator * denominator:
            best_numerator = numerator
            best_denominator = denominator
        index = end
    if best_numerator < 0:
        raise _InvalidSource("xlsx_formula_ml_threshold_search_failed")
    with localcontext() as context:
        context.prec = 50
        return Decimal(best_numerator) / Decimal(best_denominator)


def _metric_answer(path: Path, digits: int) -> str:
    from openpyxl import load_workbook

    formula_workbook = load_workbook(
        path, read_only=True, data_only=False, keep_links=False
    )
    value_workbook = load_workbook(
        path, read_only=True, data_only=True, keep_links=False
    )
    try:
        sheets = _snapshot_workbook(formula_workbook, value_workbook)
        coefficient = _coefficient_table(sheets)
        standardized = _standardized_table(sheets, coefficient)
        f1 = _maximum_f1(coefficient, standardized)
    finally:
        formula_workbook.close()
        value_workbook.close()
    quantum = Decimal(1).scaleb(-digits)
    rounded = f1.quantize(quantum, rounding=ROUND_HALF_UP)
    return format(rounded, f".{digits}f")


def _decision(
    answer: str,
    path: Path,
    root: Path,
    operations: int,
    reason: str,
) -> StructuredCandidateDecision:
    return StructuredCandidateDecision(
        "resolved",
        reason,
        StructuredCandidateAnswer(
            answer=answer,
            source_paths=(
                unicodedata.normalize("NFC", path.relative_to(root).as_posix()),
            ),
            source_sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
            operation_count=operations,
            output_count=1,
        ),
    )


def _hold(reason: str) -> StructuredCandidateDecision:
    return StructuredCandidateDecision("hold", reason)


def decide_question(engine: Any, question: str) -> StructuredCandidateDecision | None:
    contract = graph_contract_for_question(question)
    if contract is None:
        return None
    root = _source_root(engine)
    if root is None:
        return _hold("xlsx_formula_ml_source_root_invalid")
    bindings = contract["bindings"]
    paths = _named_workbooks(engine, bindings["location"], bindings["container"])
    if len(paths) != 1:
        return _hold("xlsx_formula_ml_workbook_not_unique")
    try:
        _validate_archive(paths[0])
        if contract["rule_id"] == "xlsx_yellow_residual_formula_lineage_attribute":
            answer = _lineage_answer(paths[0], bindings["attribute"])
            reason = "certified_xlsx_formula_lineage_raw_verified"
        else:
            digits = int(unicodedata.normalize("NFKC", bindings["digits"]))
            answer = _metric_answer(paths[0], digits)
            reason = "certified_xlsx_standardized_f1_argmax"
    except (
        ArithmeticError,
        KeyError,
        OSError,
        RuntimeError,
        TypeError,
        ValueError,
        zipfile.BadZipFile,
    ):
        return _hold("xlsx_formula_ml_source_invalid")
    return _decision(
        answer,
        paths[0],
        root,
        len(contract["operation_graph"]["nodes"]),
        reason,
    )


def decide_from_graph(
    engine: Any,
    question: str,
    graph_plan: Any,
) -> StructuredCandidateDecision | None:
    contract = graph_contract_for_question(question)
    if contract is None:
        return None
    if (
        graph_plan is None
        or getattr(graph_plan, "original_question", None) != question
        or getattr(graph_plan, "strict_status", None) != "pass"
    ):
        return _hold("xlsx_formula_ml_graph_plan_not_certified")
    branches = getattr(graph_plan, "branch_intents", ())
    if not isinstance(branches, tuple) or len(branches) != 1:
        return _hold("xlsx_formula_ml_graph_plan_not_certified")
    branch = branches[0]
    if not isinstance(branch, Mapping) or branch.get("status") != "resolved":
        return _hold("xlsx_formula_ml_graph_plan_not_certified")
    intent = branch.get("intent")
    supplied = (
        intent.get("extended_graph_contract") if isinstance(intent, Mapping) else None
    )
    if (
        not isinstance(supplied, Mapping)
        or not validate_graph_contract(question, supplied)
        or _canonical_json(supplied) != _canonical_json(contract)
    ):
        return _hold("xlsx_formula_ml_graph_plan_contract_mismatch")
    return decide_question(engine, question)


__all__ = [
    "REGRESSION_F1_MAX",
    "XLSX_FORMULA_ML_RULE_VERSION",
    "YELLOW_FORMULA_ATTRIBUTE",
    "decide_from_graph",
    "decide_question",
    "graph_contract_for_question",
    "validate_graph_contract",
]
