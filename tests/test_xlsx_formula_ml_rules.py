from __future__ import annotations

import copy
import math
import shutil
import sys
import tempfile
import unittest
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.styles import PatternFill


ROOT = Path(__file__).resolve().parents[1]
RAG = ROOT / "rag"
if str(RAG) not in sys.path:
    sys.path.insert(0, str(RAG))

from answer import validate_graph_answer  # noqa: E402
from question_graph_runtime import build_graph_plan  # noqa: E402
from xlsx_formula_ml_rules import (  # noqa: E402
    decide_from_graph,
    decide_question,
    graph_contract_for_question,
    validate_graph_contract,
)


S_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PR_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
S = "{" + S_NS + "}"
R = "{" + R_NS + "}"
PR = "{" + PR_NS + "}"


def engine_for(root: Path) -> SimpleNamespace:
    return SimpleNamespace(
        source_root=root.resolve(),
        glossary=SimpleNamespace(entries={}),
    )


def lineage_question(
    *,
    location: str = "架空部門",
    container: str = "sample.xlsx",
    entity: str = "建物",
) -> str:
    return (
        f"{location}の{container}において、黄色ハイライトセルは予測と実際の誤差を"
        f"計算していますが、その予測値の対象となっている{entity}の建設年を算出してください。"
    )


def metric_question(
    *, location: str = "架空分析室", container: str = "model.xlsx", digits: int = 5
) -> str:
    return (
        f"{location}の{container}にて算出された回帰係数を用いて全データの予測値を"
        "計算し、正解データに対する F1 スコアが最大となるように閾値を設定したときの "
        f"F1 スコアを答えてください。小数第{digits}位まで求めてください。"
    )


def _cached_formula_values(path: Path, values: dict[tuple[str, str], object]) -> None:
    with zipfile.ZipFile(path) as archive:
        members = {info.filename: archive.read(info) for info in archive.infolist()}
    workbook = ET.fromstring(members["xl/workbook.xml"])
    relationships = ET.fromstring(members["xl/_rels/workbook.xml.rels"])
    relation_map = {
        node.get("Id"): node.get("Target")
        for node in relationships.findall(PR + "Relationship")
    }
    parts: dict[str, str] = {}
    sheets = workbook.find(S + "sheets")
    assert sheets is not None
    for sheet in sheets.findall(S + "sheet"):
        target = relation_map[sheet.get(R + "id")]
        assert target is not None
        target = target.lstrip("/")
        parts[str(sheet.get("name"))] = (
            target if target.startswith("xl/") else "xl/" + target
        )
    for sheet_name in {sheet for sheet, _ in values}:
        root = ET.fromstring(members[parts[sheet_name]])
        cells = {
            str(cell.get("r")): cell for cell in root.findall(".//" + S + "c")
        }
        for (name, coordinate), value in values.items():
            if name != sheet_name:
                continue
            cell = cells[coordinate]
            value_node = cell.find(S + "v")
            if value_node is None:
                value_node = ET.SubElement(cell, S + "v")
            value_node.text = str(value)
        members[parts[sheet_name]] = ET.tostring(
            root, encoding="utf-8", xml_declaration=True
        )
    temporary = path.with_suffix(".rewrite.xlsx")
    with zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, data in members.items():
            archive.writestr(name, data)
    temporary.replace(path)


def write_lineage_book(
    path: Path,
    *,
    construction_year: int = 1975,
    duplicate_yellow: bool = False,
    raw_mismatch: bool = False,
    malformed_projection: bool = False,
) -> None:
    base_year = 2030
    age = base_year - construction_year
    category = 2
    actual = 800
    intercept = 1000
    age_coefficient = -3
    category_coefficient = 7
    prediction = intercept + age_coefficient * age + category_coefficient * category
    squared_error = (prediction - actual) ** 2

    workbook = Workbook()
    calc = workbook.active
    calc.title = "Calc"
    calc["A17"] = "切片"
    calc["B17"] = intercept
    calc["A18"] = "YEAR BUILT_fillna"
    calc["B18"] = age_coefficient
    calc["A19"] = "category"
    calc["B19"] = category_coefficient
    calc["B22"] = (
        "=(B18*FeatureData!F4+Calc!B19*FeatureData!G4+Calc!B17-FeatureData!H4)^2"
    )
    calc["B22"].fill = PatternFill(
        fill_type="solid", fgColor="FFFFFF00"
    )
    if duplicate_yellow:
        calc["C22"] = "=1+1"
        calc["C22"].fill = PatternFill(
            fill_type="solid", fgColor="FFFFFF00"
        )

    data = workbook.create_sheet("FeatureData")
    data.append(
        [
            "id",
            "YEAR BUILT",
            "YEAR BUILT_fillna",
            "category",
            "target",
            "YEAR BUILT_fillna",
            "category",
            "target",
        ]
    )
    for _ in range(2):
        data.append([])
    data["A4"] = "opaque-record"
    data["B4"] = construction_year
    data["C4"] = f"={base_year}-IF(B4>0,B4,1990)"
    data["D4"] = category
    data["E4"] = actual
    data["F4"] = (
        "=C4" if malformed_projection else
        "=INDEX($A4:$E4,1,MATCH(F$1,$A$1:$E$1,0))"
    )
    data["G4"] = "=INDEX($A4:$E4,1,MATCH(G$1,$A$1:$E$1,0))"
    data["H4"] = "=INDEX($A4:$E4,1,MATCH(H$1,$A$1:$E$1,0))"

    raw = workbook.create_sheet("Ledger")
    raw.append(["id", "YEAR BUILT", "category", "target"])
    raw.append(
        [
            "opaque-record",
            construction_year + (1 if raw_mismatch else 0),
            category,
            actual,
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    _cached_formula_values(
        path,
        {
            ("Calc", "B22"): squared_error,
            ("FeatureData", "C4"): age,
            ("FeatureData", "F4"): age,
            ("FeatureData", "G4"): category,
            ("FeatureData", "H4"): actual,
            **({("Calc", "C22"): 2} if duplicate_yellow else {}),
        },
    )


def _population_standardize(values: list[float]) -> list[float]:
    mean = math.fsum(values) / len(values)
    deviation = math.sqrt(
        math.fsum((value - mean) ** 2 for value in values) / len(values)
    )
    return [(value - mean) / deviation for value in values]


def write_metric_book(
    path: Path,
    *,
    labels: tuple[int, ...] = (1, 1, 0, 0, 0, 0),
    stale_standardized: bool = False,
    duplicate_coefficients: bool = False,
    raw_mismatch: bool = False,
    duplicate_raw: bool = False,
) -> None:
    first = [6.0, 5.0, 5.0, 4.0, 3.0, 2.0]
    second = [10.0, 11.0, 12.0, 13.0, 14.0, 15.0]
    assert len(labels) == len(first)
    standardized_first = _population_standardize(first)
    standardized_second = _population_standardize(second)
    last_row = len(first) + 1

    workbook = Workbook()
    model = workbook.active
    model.title = "Model"
    model["A1"] = "概要"
    model["A8"] = "観測数"
    model["B8"] = len(first)
    model["B16"] = "係数"
    model["A17"] = "切片"
    model["B17"] = 0
    model["A18"] = "signal_a"
    model["B18"] = 1
    model["A19"] = "signal_b"
    model["B19"] = 0

    scaled = workbook.create_sheet("Scaled")
    scaled.append(["signal_a", "signal_b", "truth"])
    extracted = workbook.create_sheet("Extract")
    extracted.append(["signal_a", "signal_b", "truth"])
    raw = workbook.create_sheet("Raw")
    raw.append(["id", "signal_a", "signal_b", "truth"])
    cached: dict[tuple[str, str], object] = {}
    for index, (left, right, target) in enumerate(
        zip(first, second, labels), start=2
    ):
        scaled.cell(index, 1).value = (
            f"=STANDARDIZE(Extract!A{index},AVERAGE(Extract!A$2:A${last_row}),"
            f"_xlfn.STDEV.P(Extract!A$2:A${last_row}))"
        )
        scaled.cell(index, 2).value = (
            f"=STANDARDIZE(Extract!B{index},AVERAGE(Extract!B$2:B${last_row}),"
            f"_xlfn.STDEV.P(Extract!B$2:B${last_row}))"
        )
        scaled.cell(index, 3).value = target
        extracted.append([left, right, target])
        raw.append(
            [
                f"record-{index - 1}",
                left + (1 if raw_mismatch and index == 2 else 0),
                right,
                target,
            ]
        )
        cached[("Scaled", f"A{index}")] = (
            standardized_first[index - 2]
            + (0.25 if stale_standardized and index == 2 else 0)
        )
        cached[("Scaled", f"B{index}")] = standardized_second[index - 2]
    if duplicate_coefficients:
        copy_sheet = workbook.create_sheet("ModelCopy")
        copy_sheet["B4"] = "係数"
        copy_sheet["A5"] = "切片"
        copy_sheet["B5"] = 0
        copy_sheet["A6"] = "signal_a"
        copy_sheet["B6"] = 1
    if duplicate_raw:
        raw_copy = workbook.copy_worksheet(raw)
        raw_copy.title = "RawCopy"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    _cached_formula_values(path, cached)


class XlsxFormulaMlRulesTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_complete_grammars_compile_deterministically(self) -> None:
        for question in (lineage_question(), metric_question()):
            contract = graph_contract_for_question(question)
            self.assertIsNotNone(contract)
            assert contract is not None
            self.assertTrue(contract["graph_contract_id"].startswith("xlsx_formula_ml_"))
            self.assertTrue(validate_graph_contract(question, contract))
            mutated = copy.deepcopy(contract)
            mutated["bindings"]["location"] += "-changed"
            self.assertFalse(validate_graph_contract(question, mutated))
        self.assertIsNone(
            graph_contract_for_question("表計算ファイルを見て答えてください。")
        )

    def test_formula_lineage_entity_is_semantically_bound(self) -> None:
        for entity in ("不動産", "物件", "建物", "建築物", "家屋"):
            self.assertIsNotNone(
                graph_contract_for_question(lineage_question(entity=entity)), entity
            )
        self.assertIsNone(
            graph_contract_for_question(lineage_question(entity="宇宙船"))
        )

    def test_formula_lineage_resolves_and_changes_with_source(self) -> None:
        source = self.root / "架空部門" / "sample.xlsx"
        write_lineage_book(source, construction_year=1975)
        decision = decide_question(engine_for(self.root), lineage_question())
        self.assertEqual("resolved", decision.status)
        self.assertEqual("1975", decision.result.answer)
        write_lineage_book(source, construction_year=1988)
        changed = decide_question(engine_for(self.root), lineage_question())
        self.assertEqual("resolved", changed.status)
        self.assertEqual("1988", changed.result.answer)

    def test_formula_lineage_fails_closed_on_ambiguity_or_mismatch(self) -> None:
        source = self.root / "架空部門" / "sample.xlsx"
        for options in (
            {"duplicate_yellow": True},
            {"raw_mismatch": True},
            {"malformed_projection": True},
        ):
            write_lineage_book(source, **options)
            decision = decide_question(engine_for(self.root), lineage_question())
            self.assertEqual("hold", decision.status, options)

    def test_f1_search_groups_equal_scores_and_rounds_only_final_metric(self) -> None:
        source = self.root / "架空分析室" / "model.xlsx"
        # The second and third records have equal predictions.  Splitting the
        # tie would falsely produce F1=1.0; a real threshold must include both.
        write_metric_book(source, labels=(1, 1, 0, 0, 0, 0))
        decision = decide_question(engine_for(self.root), metric_question())
        self.assertEqual("resolved", decision.status)
        self.assertEqual("0.80000", decision.result.answer)

        write_metric_book(source, labels=(1, 1, 0, 1, 0, 0))
        changed = decide_question(engine_for(self.root), metric_question())
        self.assertEqual("resolved", changed.status)
        self.assertEqual("0.85714", changed.result.answer)

    def test_f1_rule_fails_closed_on_stale_or_ambiguous_source(self) -> None:
        source = self.root / "架空分析室" / "model.xlsx"
        for options in (
            {"stale_standardized": True},
            {"duplicate_coefficients": True},
            {"raw_mismatch": True},
            {"duplicate_raw": True},
        ):
            write_metric_book(source, **options)
            decision = decide_question(engine_for(self.root), metric_question())
            self.assertEqual("hold", decision.status, options)

    def test_live_graph_plan_is_mandatory_and_answer_contract_accepts_output(self) -> None:
        source = self.root / "架空分析室" / "model.xlsx"
        write_metric_book(source)
        question = metric_question()
        plan = build_graph_plan("opaque-f1", question, fast_advisory=True)
        self.assertEqual("pass", plan.strict_status)
        decision = decide_from_graph(engine_for(self.root), question, plan)
        self.assertEqual("resolved", decision.status)
        self.assertEqual((), validate_graph_answer(decision.result.answer, plan))

        no_plan = decide_from_graph(engine_for(self.root), question, None)
        self.assertEqual("hold", no_plan.status)
        branches = copy.deepcopy(plan.branch_intents)
        branches[0]["intent"]["extended_graph_contract"]["bindings"][
            "digits"
        ] = "4"
        tampered = replace(plan, branch_intents=branches)
        mismatch = decide_from_graph(engine_for(self.root), question, tampered)
        self.assertEqual("hold", mismatch.status)

    def test_duplicate_matching_workbooks_hold(self) -> None:
        source = self.root / "架空分析室" / "model.xlsx"
        write_metric_book(source)
        duplicate = self.root / "架空分析室" / "nested" / "model.xlsx"
        duplicate.parent.mkdir(parents=True)
        shutil.copy2(source, duplicate)
        decision = decide_question(engine_for(self.root), metric_question())
        self.assertEqual("hold", decision.status)


if __name__ == "__main__":
    unittest.main()
