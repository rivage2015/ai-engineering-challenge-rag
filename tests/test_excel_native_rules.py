from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
RAG = ROOT / "rag"
if str(RAG) not in sys.path:
    sys.path.insert(0, str(RAG))

from excel_native_rules import (  # noqa: E402
    decide_question,
    graph_contract_for_question,
    validate_graph_contract,
)
from glossary import Glossary  # noqa: E402
from structured_candidate import StructuredCandidateEngine  # noqa: E402


def write_conditional_book(
    path: Path,
    threshold: str,
    *,
    duplicate_visible: bool = False,
    duplicate_rule: bool = False,
    unsupported_rule: bool = False,
    visible_argb: str = "FFFFEB9C",
    visible_semantic: str = "相関",
) -> None:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import PatternFill

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    visible = workbook.active
    visible.title = "visible_corr"
    visible["A1"] = visible_semantic
    visible["A2"] = -0.5
    yellow = PatternFill(fill_type="solid", fgColor=visible_argb)
    visible.conditional_formatting.add(
        "A1:C3",
        CellIsRule(operator="lessThan", formula=[threshold], fill=yellow),
    )
    if duplicate_rule:
        visible.conditional_formatting.add(
            "D1:D3",
            CellIsRule(operator="lessThan", formula=[threshold], fill=yellow),
        )
    if unsupported_rule:
        visible.conditional_formatting.add(
            "E1:E3",
            FormulaRule(formula=["E1<0"], fill=yellow),
        )
    hidden = workbook.create_sheet("hidden_corr")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "相関"
    hidden_yellow = PatternFill(fill_type="solid", fgColor="FFFFEB9C")
    hidden.conditional_formatting.add(
        "A1:C3",
        CellIsRule(operator="lessThan", formula=["-0.9"], fill=hidden_yellow),
    )
    if duplicate_visible:
        duplicate = workbook.create_sheet("visible_corr_copy")
        duplicate["A1"] = "相関"
        duplicate.conditional_formatting.add(
            "A1:C3",
            CellIsRule(operator="lessThan", formula=[threshold], fill=yellow),
        )
    workbook.save(path)
    workbook.close()


def write_regression_book(
    path: Path,
    *,
    alpha: float = 0.25,
    duplicate_id: bool = False,
    duplicate_header: bool = False,
    duplicate_coefficient: bool = False,
    second_data_table: bool = False,
) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    analysis = workbook.active
    analysis.title = "analysis"
    analysis["B3"] = "係数"
    coefficient_rows = [("alpha", alpha), ("beta", -0.1)]
    if duplicate_coefficient:
        coefficient_rows.append(("ＡＬＰＨＡ", 0.75))
    coefficient_rows.append(("切片", 0.5))
    for row, (label, value) in enumerate(
        coefficient_rows,
        4,
    ):
        analysis.cell(row, 1, label)
        analysis.cell(row, 2, value)
    data = workbook.create_sheet("records")
    if duplicate_header:
        data.append(["row_id", "alpha", "ＡＬＰＨＡ", "beta"])
        data.append([7, 2, 9, 3])
    else:
        data.append(["row_id", "alpha", "beta", "ignored_target"])
        data.append([7, 2, 3, 99])
    if duplicate_id:
        data.append([7, 9, 9, 0])
    if second_data_table:
        second = workbook.create_sheet("records_copy")
        second.append(["row_id", "alpha", "beta"])
        second.append([8, 4, 5])
    workbook.save(path)
    workbook.close()


class ExcelNativeRulesTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.engine = StructuredCandidateEngine(self.root, Glossary())

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_visible_conditional_format_is_source_driven_and_hidden_decoy_is_ignored(self) -> None:
        source = self.root / "opaque_project" / "03.データ" / "book.xlsx"
        question = (
            "opaque_projectのbook.xlsxにおいて、表示されている相関係数シートで、"
            "黄色ハイライトになっているセルの条件を答えてください。"
        )
        contract = graph_contract_for_question(question)
        self.assertIsNotNone(contract)
        self.assertEqual(
            contract["rule_id"],
            "excel_visible_conditional_format_predicate",
        )
        self.assertTrue(validate_graph_contract(question, contract))

        write_conditional_book(source, "-0.42")
        first = decide_question(self.engine, question)
        self.assertEqual(first.status, "resolved")
        self.assertEqual(first.result.answer, "相関係数が-0.42未満")
        first_sha = first.result.source_sha256

        write_conditional_book(source, "-0.55")
        changed = decide_question(self.engine, question)
        self.assertEqual(changed.status, "resolved")
        self.assertEqual(changed.result.answer, "相関係数が-0.55未満")
        self.assertNotEqual(changed.result.source_sha256, first_sha)

        write_conditional_book(source, "-0.55", duplicate_visible=True)
        ambiguous = decide_question(self.engine, question)
        self.assertEqual(ambiguous.status, "hold")

    def test_visible_conditional_format_holds_on_rule_or_semantic_ambiguity(self) -> None:
        source = self.root / "opaque_project" / "03.データ" / "book.xlsx"
        question = (
            "opaque_projectのbook.xlsxにおいて、表示されている相関係数シートで、"
            "黄色ハイライトになっているセルの条件を答えてください。"
        )
        cases = (
            {"duplicate_rule": True},
            {"unsupported_rule": True},
            {"visible_argb": "FFFFF2CC"},
            {"visible_semantic": "相関行列"},
        )
        for options in cases:
            with self.subTest(options=options):
                write_conditional_book(source, "-0.55", **options)
                decision = decide_question(self.engine, question)
                self.assertEqual(decision.status, "hold")

    def test_raw_regression_coefficients_align_fields_and_round_only_final_value(self) -> None:
        source = self.root / "opaque_project" / "03.データ" / "book.xlsx"
        question = (
            "opaque_projectのbook.xlsxにおいて、回帰分析の結果として記載されている係数を"
            "row_id=7のデータに当てはめたときの予測値はいくつですか。"
            "小数第3位まで答えてください。"
        )
        contract = graph_contract_for_question(question)
        self.assertIsNotNone(contract)
        self.assertEqual(
            contract["rule_id"],
            "excel_raw_regression_coefficient_prediction",
        )
        self.assertEqual(
            contract["requested_output"]["display_precision"],
            {"mode": "decimal_places", "digits": 3},
        )
        self.assertTrue(validate_graph_contract(question, contract))

        write_regression_book(source)
        first = decide_question(self.engine, question)
        self.assertEqual(first.status, "resolved")
        self.assertEqual(first.result.answer, "0.700")

        write_regression_book(source, alpha=0.5)
        changed = decide_question(self.engine, question)
        self.assertEqual(changed.status, "resolved")
        self.assertEqual(changed.result.answer, "1.200")

        write_regression_book(source, duplicate_id=True)
        ambiguous = decide_question(self.engine, question)
        self.assertEqual(ambiguous.status, "hold")

    def test_raw_regression_holds_on_coefficient_field_table_or_row_ambiguity(self) -> None:
        source = self.root / "opaque_project" / "03.データ" / "book.xlsx"
        question = (
            "opaque_projectのbook.xlsxにおいて、回帰分析の結果として記載されている係数を"
            "row_id=7のデータに当てはめたときの予測値はいくつですか。"
            "小数第3位まで答えてください。"
        )
        cases = (
            {"duplicate_header": True},
            {"duplicate_coefficient": True},
            {"second_data_table": True},
            {"duplicate_id": True},
        )
        for options in cases:
            with self.subTest(options=options):
                write_regression_book(source, **options)
                decision = decide_question(self.engine, question)
                self.assertEqual(decision.status, "hold")


if __name__ == "__main__":
    unittest.main()
